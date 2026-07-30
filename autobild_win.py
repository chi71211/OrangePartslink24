"""
AutoBild 爬蟲系統 v11.0 - Windows 版本
========================================
用法：
  python autobild_win.py              # 完整掃描
  python autobild_win.py --test       # 測試模式 (2 廠牌 x 2 車系)
  python autobild_win.py --reset      # 重置資料庫
  python autobild_win.py --status     # 查看統計
  python autobild_win.py --brand VW   # 只抓特定品牌
"""

import subprocess
import sys

def check_packages():
    """檢查必要套件是否已安裝"""
    required = {'nest_asyncio': 'nest_asyncio', 'playwright': 'playwright', 'pandas': 'pandas', 'openpyxl': 'openpyxl'}
    missing = []
    for pkg, pip_name in required.items():
        try:
            __import__(pkg)
        except ImportError:
            missing.append(pip_name)
    
    if missing:
        print("[Error] Missing packages:")
        for p in missing:
            print(f"  - {p}")
        print("\nPlease install manually:")
        print(f"  pip install {' '.join(missing)}")
        sys.exit(1)

check_packages()

import nest_asyncio
nest_asyncio.apply()
import asyncio
import os
import glob
import sys
import time
import random
import sqlite3
import argparse
import pandas as pd
import re
import signal
from datetime import datetime, timedelta
from urllib.parse import urljoin
from playwright.async_api import async_playwright
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font

os.chdir(os.path.dirname(os.path.abspath(__file__)))


class Config:
    BASE_URL = "https://www.autobild.de"
    CATALOG_URL = f"{BASE_URL}/marken-modelle/#aktuell"
    CSV_DIR = "AutoBild_Exports"
    DB_FILE = "autobild_master.db"
    BATCH_SIZE = 20
    DELAY_MIN = 0.8
    DELAY_MAX = 1.6
    MAX_RETRIES = 3
    MAX_RUNTIME_HOURS = 5.5
    API_WAIT_TIMEOUT = 8
    SNAPSHOT_FILE = os.path.join(CSV_DIR, "autobild_snapshot.xlsx")
    DIFF_FILE = os.path.join(CSV_DIR, "autobild_diff.xlsx")


class DatabaseManager:
    def __init__(self):
        os.makedirs(Config.CSV_DIR, exist_ok=True)
        self.conn = sqlite3.connect(Config.DB_FILE)
        self.cursor = self.conn.cursor()
        self.batch = []
        self._init_db()
        self.changelog_path = Config.CSV_DIR + '/autobild_changelog.xlsx'
        self._init_changelog()

    def _init_changelog(self):
        try:
            if os.path.exists(self.changelog_path):
                wb = load_workbook(self.changelog_path)
                if 'Changelog' not in wb.sheetnames:
                    ws = wb.create_sheet('Changelog', 0)
                    ws.append(['Timestamp', 'Action', 'Brand', 'Model', 'Category', 'Fuel_Type', 'Typ', 'Start_Year', 'End_Year', 'HSN_TSN_Old', 'HSN_TSN_New'])
                    ws.column_dimensions['A'].width = 20
                    for col in ['B','C','D','E','F','G','H','I','J','K']:
                        ws.column_dimensions[col].width = 16
                wb.close()
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = 'Changelog'
                ws.append(['Timestamp', 'Action', 'Brand', 'Model', 'Category', 'Fuel_Type', 'Typ', 'Start_Year', 'End_Year', 'HSN_TSN_Old', 'HSN_TSN_New'])
                ws.column_dimensions['A'].width = 20
                for col in ['B','C','D','E','F','G','H','I','J','K']:
                    ws.column_dimensions[col].width = 16
                wb.save(self.changelog_path)
                wb.close()
        except Exception:
            pass

    def _log_changelog(self, records):
        try:
            wb = load_workbook(self.changelog_path)
            ws = wb['Changelog']
            green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            for action, r in records:
                ts = time.strftime('%Y-%m-%d %H:%M:%S')
                row_data = [ts, action, r.get('Brand',''), r.get('Model',''), r.get('Category',''), r.get('Fuel_Type',''), r.get('Typ',''), r.get('Start_Year',''), r.get('End_Year',''), r.get('_hsn_old',''), r.get('HSN_TSN','')]
                ws.append(row_data)
                new_row = ws.max_row
                fill = green if action == 'INSERT' else red
                for col in range(1, len(row_data) + 1):
                    ws.cell(row=new_row, column=col).fill = fill
            wb.save(self.changelog_path)
            wb.close()
        except Exception:
            pass

    def _init_db(self):
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS car_catalog (
                Brand TEXT,
                Model TEXT,
                Category TEXT,
                Fuel_Type TEXT,
                Typ TEXT,
                Start_Year TEXT,
                End_Year TEXT,
                HSN_TSN TEXT,
                UNIQUE(Brand, Model, Category, Fuel_Type, Typ, Start_Year, End_Year, HSN_TSN)
            )
        ''')

        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS model_progress (
                Brand TEXT,
                Model TEXT,
                variant_count INTEGER,
                last_scraped TEXT,
                PRIMARY KEY(Brand, Model)
            )
        ''')

        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS system_metadata (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        ''')

        self.cursor.execute('DROP VIEW IF EXISTS view_car_catalog')
        self.cursor.execute('''
            CREATE VIEW view_car_catalog AS
            SELECT
                Brand, Model, Category, Fuel_Type, Typ, Start_Year, End_Year, HSN_TSN
            FROM car_catalog
            GROUP BY Brand, Model, Category, Fuel_Type, Typ, Start_Year, End_Year, HSN_TSN
        ''')
        self.conn.commit()

    def add_to_batch(self, record: dict):
        self.batch.append(record)
        if len(self.batch) >= Config.BATCH_SIZE:
            self.flush()

    def flush(self):
        if not self.batch:
            return
        changelog_entries = []
        for r in self.batch:
            brand = r.get('Brand', 'N/A')
            model = r.get('Model', 'N/A')
            cat = r.get('Category', 'N/A')
            fuel = r.get('Fuel_Type', 'N/A')
            typ = r.get('Typ', 'N/A')
            sy = r.get('Start_Year', 'N/A')
            ey = r.get('End_Year', 'N/A')
            hsn = r.get('HSN_TSN', 'N/A')

            self.cursor.execute('''
                SELECT rowid, HSN_TSN FROM car_catalog
                WHERE Brand=? AND Model=? AND Category=? AND Fuel_Type=? AND Typ=?
                  AND Start_Year=? AND End_Year=?
            ''', (brand, model, cat, fuel, typ, sy, ey))
            existing = self.cursor.fetchone()

            if existing:
                rowid, old_hsn = existing
                if old_hsn != hsn:
                    self.cursor.execute('UPDATE car_catalog SET HSN_TSN=? WHERE rowid=?', (hsn, rowid))
                    r['_hsn_old'] = old_hsn
                    changelog_entries.append(('UPDATE', r))
                else:
                    pass
            else:
                self.cursor.execute('''
                    INSERT INTO car_catalog
                    (Brand, Model, Category, Fuel_Type, Typ, Start_Year, End_Year, HSN_TSN)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (brand, model, cat, fuel, typ, sy, ey, hsn))
                r['_hsn_old'] = ''
                changelog_entries.append(('INSERT', r))
        self.conn.commit()
        if changelog_entries:
            self._log_changelog(changelog_entries)
        self.batch.clear()

    def get_progress(self, brand: str, model: str):
        self.cursor.execute(
            'SELECT variant_count, last_scraped FROM model_progress WHERE Brand=? AND Model=?',
            (brand, model)
        )
        row = self.cursor.fetchone()
        if not row:
            return None
        count, last_scraped = row
        try:
            scraped_dt = datetime.fromisoformat(last_scraped)
            if datetime.now() - scraped_dt > timedelta(days=7):
                return None
        except Exception:
            pass
        return count

    def update_progress(self, brand: str, model: str, count: int):
        self.cursor.execute('''
            INSERT OR REPLACE INTO model_progress (Brand, Model, variant_count, last_scraped)
            VALUES (?, ?, ?, ?)
        ''', (brand, model, count, datetime.now().isoformat()))
        self.conn.commit()

    def mark_brand_done(self, brand: str):
        self.cursor.execute("REPLACE INTO system_metadata (key, value) VALUES (?, 'true')",
                           (f'brand_done_{brand}',))
        self.conn.commit()

    def is_brand_done(self, brand: str):
        self.cursor.execute("SELECT value FROM system_metadata WHERE key=?", (f'brand_done_{brand}',))
        row = self.cursor.fetchone()
        return row is not None and row[0] == 'true'

    def get_all_data(self):
        self.flush()
        try:
            return pd.read_sql_query("SELECT * FROM view_car_catalog ORDER BY Brand, Model, Fuel_Type, Typ", self.conn)
        except Exception:
            return pd.DataFrame()

    def export_brand_csv(self, brand: str):
        self.flush()
        try:
            df = pd.read_sql_query(
                "SELECT * FROM view_car_catalog WHERE Brand = ?",
                self.conn, params=(brand,)
            )
            if not df.empty:
                path = os.path.join(Config.CSV_DIR, f"{brand}.csv")
                df.to_csv(path, index=False, encoding='utf-8-sig')
                return len(df)
        except Exception:
            pass
        return 0

    def get_stats(self):
        try:
            df = pd.read_sql_query('''
                SELECT Brand, Fuel_Type, Category,
                       COUNT(DISTINCT Model) as Models,
                       COUNT(*) as Rows
                FROM view_car_catalog
                GROUP BY Brand, Fuel_Type, Category
                ORDER BY Brand, Fuel_Type
            ''', self.conn)
            return df
        except Exception:
            return pd.DataFrame()

    def close(self):
        self.flush()
        self.conn.close()


class SnapshotManager:
    SNAPSHOT_FILE = Config.SNAPSHOT_FILE
    DIFF_FILE = Config.DIFF_FILE
    KEY_COLS = ['Brand', 'Model', 'Fuel_Type', 'Typ']
    ALL_COLS = ['Brand', 'Model', 'Category', 'Fuel_Type', 'Typ', 'Start_Year', 'End_Year', 'HSN_TSN']

    RED_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    GREEN_FILL = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    GRAY_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_FONT = Font(color='FFFFFF', bold=True)
    COMPARE_FIELDS = ['Category', 'Start_Year', 'End_Year', 'HSN_TSN']

    @staticmethod
    def load_snapshot():
        try:
            df = pd.read_excel(SnapshotManager.SNAPSHOT_FILE, engine='openpyxl')
            print(f"  [Snapshot] Loaded previous snapshot ({len(df)} records)")
            return df
        except Exception:
            return None

    @staticmethod
    def save_snapshot(df):
        try:
            today = datetime.now().strftime('%Y-%m-%d')
            if os.path.exists(SnapshotManager.SNAPSHOT_FILE):
                bak = SnapshotManager.SNAPSHOT_FILE.replace('.xlsx', f'_{today}.xlsx')
                if not os.path.exists(bak):
                    os.rename(SnapshotManager.SNAPSHOT_FILE, bak)
            out = df[SnapshotManager.ALL_COLS] if all(c in df.columns for c in SnapshotManager.ALL_COLS) else df
            out.to_excel(SnapshotManager.SNAPSHOT_FILE, index=False, engine='openpyxl')
        except Exception as e:
            print(f"  [Snapshot] Save failed: {e}")

    @staticmethod
    def generate_diff_report(old_df, new_df):
        if old_df is None or old_df.empty:
            SnapshotManager.save_snapshot(new_df)
            print("  [Diff] First run, snapshot created")
            return

        for col in SnapshotManager.ALL_COLS:
            if col in old_df.columns:
                old_df[col] = old_df[col].apply(lambda x: '' if pd.isna(x) else str(x).strip())
            if col in new_df.columns:
                new_df[col] = new_df[col].apply(lambda x: '' if pd.isna(x) else str(x).strip())

        old_df.replace(to_replace=r'^(N/A|n/a|NA|na|nan|NaN|Nan|null|None)$', value='', regex=True, inplace=True)
        new_df.replace(to_replace=r'^(N/A|n/a|NA|na|nan|NaN|Nan|null|None)$', value='', regex=True, inplace=True)

        missing = [c for c in SnapshotManager.KEY_COLS if c not in old_df.columns or c not in new_df.columns]
        if missing:
            print(f"  [Diff] Missing key columns {missing}, skipping")
            SnapshotManager.save_snapshot(new_df)
            return

        old_df['_key'] = old_df[SnapshotManager.KEY_COLS].agg('|'.join, axis=1)
        new_df['_key'] = new_df[SnapshotManager.KEY_COLS].agg('|'.join, axis=1)

        old_df = old_df.drop_duplicates(subset='_key', keep='last')
        new_df = new_df.drop_duplicates(subset='_key', keep='last')

        old_keys = set(old_df['_key'])
        new_keys = set(new_df['_key'])

        new_records = new_df[~new_df['_key'].isin(old_keys)].copy()
        removed_records = old_df[~old_df['_key'].isin(new_keys)].copy()

        common_keys = old_keys & new_keys
        old_idx = old_df[old_df['_key'].isin(common_keys)].set_index('_key')
        new_idx = new_df[new_df['_key'].isin(common_keys)].set_index('_key')

        compare_cols = [c for c in SnapshotManager.ALL_COLS if c not in SnapshotManager.KEY_COLS]

        modified = []
        for key in common_keys:
            old_row = old_idx.loc[key]
            new_row = new_idx.loc[key]
            changes = {}
            for col in compare_cols:
                ov = str(old_row.get(col, ''))
                nv = str(new_row.get(col, ''))
                if ov != nv:
                    changes[col] = (ov, nv)
            if changes:
                modified.append((old_row.to_dict(), new_row.to_dict(), changes))

        has_new = not new_records.empty
        has_mod = len(modified) > 0
        has_del = not removed_records.empty

        if not (has_new or has_mod or has_del):
            print("  [Diff] No differences found")
            SnapshotManager.save_snapshot(new_df)
            return

        today = datetime.now().strftime('%Y-%m-%d')
        if os.path.exists(SnapshotManager.DIFF_FILE):
            wb = load_workbook(SnapshotManager.DIFF_FILE)
            if today in wb.sheetnames:
                ws = wb[today]
                row = ws.max_row + 2
                ws.cell(row=row - 1, column=1, value=f'--- {datetime.now().strftime("%H:%M")} ---')
            else:
                ws = wb.create_sheet(today, 0)
                row = 1
        else:
            wb = Workbook()
            wb.remove(wb.active)
            ws = wb.create_sheet(today, 0)
            row = 1

        if row == 1:
            headers = ['Status'] + SnapshotManager.KEY_COLS
            for f in SnapshotManager.COMPARE_FIELDS:
                headers.append(f'{f}(O)')
                headers.append(f'{f}(N)')
            SnapshotManager._write_header(ws, headers)
            row = 2
        for _, old_row in removed_records.iterrows():
            SnapshotManager._write_status_row(ws, row, 'Removed', old_row.to_dict(), None, SnapshotManager.RED_FILL, None)
            row += 1
        for old_data, new_data, changes in modified:
            SnapshotManager._write_status_row(ws, row, 'Modified', old_data, new_data, SnapshotManager.RED_FILL, SnapshotManager.GREEN_FILL, changes)
            row += 1
        for _, new_row in new_records.iterrows():
            SnapshotManager._write_status_row(ws, row, 'New', None, new_row.to_dict(), None, SnapshotManager.GREEN_FILL)
            row += 1

        wb.save(SnapshotManager.DIFF_FILE)
        print(f"\n  [Diff] Report: {SnapshotManager.DIFF_FILE} (tab: {today})")
        print(f"    New: {len(new_records)} | Modified: {len(modified)} | Removed: {len(removed_records)}")
        SnapshotManager.save_snapshot(new_df)

    @staticmethod
    def _write_header(ws, headers):
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = SnapshotManager.HEADER_FILL
            cell.font = SnapshotManager.HEADER_FONT

    @staticmethod
    def _write_status_row(ws, row, status, old_data, new_data, old_fill, new_fill, changes=None):
        ws.cell(row=row, column=1, value=status)
        for ci, col in enumerate(SnapshotManager.KEY_COLS, 2):
            val = (new_data or old_data or {}).get(col, '')
            ws.cell(row=row, column=ci, value=val)
        for fi, f in enumerate(SnapshotManager.COMPARE_FIELDS):
            col_o = 2 + len(SnapshotManager.KEY_COLS) + fi * 2
            col_n = col_o + 1
            ov = (old_data or {}).get(f, '')
            nv = (new_data or {}).get(f, '')
            ws.cell(row=row, column=col_o, value=ov if ov else '')
            ws.cell(row=row, column=col_n, value=nv if nv else '')
            is_changed = changes is None or f in changes
            if old_fill and ov and is_changed: ws.cell(row=row, column=col_o).fill = old_fill
            if new_fill and nv and is_changed: ws.cell(row=row, column=col_n).fill = new_fill


async def smart_delay(success=True):
    delay = random.uniform(Config.DELAY_MIN, Config.DELAY_MAX) if success else random.uniform(2.5, 4.5)
    if random.random() < 0.12:
        delay += random.uniform(1.0, 2.5)
    await asyncio.sleep(max(0.5, min(delay, 6.0)))


async def dismiss_cookie(page):
    try:
        iframe = page.frame_locator('iframe[id^="sp_message_iframe"]')
        btn = iframe.get_by_role("button", name="Alle akzeptieren")
        await btn.click(timeout=5000)
    except Exception:
        pass


def clean_text(text):
    if not text:
        return "N/A"
    t = re.sub(r'\s+', ' ', text).strip()
    t = t.rstrip(':').strip()
    if not t or t == '-' or t.lower() == 'n/a':
        return "N/A"
    return t


def extract_date_range(text):
    match = re.search(r'(\d{2}/\d{4})\s*[–-]\s*(\d{2}/\d{4})', text)
    if match:
        return f"{match.group(1)} - {match.group(2)}"
    match = re.search(r'seit\s+(\d{2}/\d{4})', text)
    if match:
        return f"seit {match.group(1)}"
    return "N/A"


class AutoBildScraper:
    def __init__(self, test_mode=False, target_brand=None, diff=True):
        self.test_mode = test_mode
        self.target_brand = target_brand.upper() if target_brand else None
        self.diff = diff
        self.db = DatabaseManager()
        self.start_time = time.time()
        self.api_hsn_tsn = None
        self.api_captured = False
        self.is_interrupted = False
        self._setup_signal_handlers()

    def _setup_signal_handlers(self):
        def signal_handler(signum, frame):
            print(f"\n\n[Warning] Interrupted (Ctrl+C), safely saving progress...")
            self.is_interrupted = True
        try:
            signal.signal(signal.SIGINT, signal_handler)
            signal.signal(signal.SIGTERM, signal_handler)
        except Exception:
            pass

    def should_stop(self):
        if self.is_interrupted:
            return True
        elapsed = (time.time() - self.start_time) / 3600
        if elapsed >= Config.MAX_RUNTIME_HOURS:
            print(f"\n[Timeout] Protection triggered (>{Config.MAX_RUNTIME_HOURS} hours), stopping...")
            return True
        return False

    def is_timeout(self):
        elapsed = (time.time() - self.start_time) / 3600
        return elapsed >= Config.MAX_RUNTIME_HOURS

    async def handle_api_response(self, response):
        try:
            url = response.url.lower()
            if '/api/vehicle-details/' in url and response.status == 200:
                ct = response.headers.get('content-type', '')
                if 'json' in ct or url.endswith('.json'):
                    try:
                        data = await response.json()
                        if isinstance(data, (dict, list)):
                            self.api_hsn_tsn = data
                            self.api_captured = True
                    except Exception:
                        pass
        except Exception:
            pass

    async def collect_brand_urls(self, page):
        print("\n[Step 1] Collecting brand URLs...")
        for attempt in range(3):
            try:
                await page.goto(Config.CATALOG_URL, timeout=90000, wait_until="domcontentloaded")
                break
            except Exception as e:
                if attempt < 2:
                    wait = (attempt + 1) * 5
                    print(f"    [Retry] 品牌列表載入失敗 ({e.__class__.__name__}), {wait}秒後重試 ({attempt+1}/3)...")
                    await asyncio.sleep(wait)
                else:
                    raise
        await dismiss_cookie(page)

        for scroll_y in range(0, 3000, 600):
            await page.evaluate(f"window.scrollTo(0, {scroll_y})")
            await asyncio.sleep(0.5)
        await smart_delay()

        raw_links = await page.evaluate(r'''() => {
            return Array.from(document.querySelectorAll('a[href*="/marken-modelle/"]'))
                .map(a => ({
                    href: a.getAttribute('href'),
                    text: (a.textContent || '').trim()
                }))
                .filter(l => l.href && l.text.length > 0);
        }''')

        brand_urls = []
        seen = set()
        for item in raw_links:
            href = item['href']
            if not href or '/marken-modelle/' not in href:
                continue

            full = urljoin(Config.BASE_URL, href).split('#')[0].split('?')[0]
            full = full.rstrip('/') + '/'

            parts = full.replace(Config.BASE_URL, '').strip('/').split('/')
            if len(parts) == 2 and parts[0] == 'marken-modelle' and parts[1]:
                if full not in seen:
                    seen.add(full)
                    brand_urls.append(full)

        brand_urls.sort(reverse=True)
        if self.target_brand:
            brand_urls = [u for u in brand_urls if f'/{self.target_brand.lower()}/' in u]
        elif self.test_mode:
            brand_urls = brand_urls[:2]

        print(f"   Found {len(brand_urls)} brands")
        return brand_urls

    async def collect_model_urls(self, page, brand_url):
        brand_name = brand_url.strip('/').split('/')[-1].upper()
        print(f"\n> Entering brand: {brand_name}")
        for attempt in range(3):
            try:
                await page.goto(brand_url, timeout=90000, wait_until="domcontentloaded")
                break
            except Exception as e:
                if 'ERR_NETWORK_CHANGED' in str(e) and attempt < 2:
                    wait = (attempt + 1) * 3
                    print(f"    [Retry] 網路不穩 ({e.__class__.__name__}), {wait}秒後重試 ({attempt+1}/3)...")
                    await asyncio.sleep(wait)
                else:
                    raise
        await dismiss_cookie(page)
        await smart_delay()

        await page.evaluate(r'''() => {
            const allEls = Array.from(document.querySelectorAll('a, button, div, span, li, h2, h3'));
            const btn = allEls.find(el => {
                const txt = (el.textContent || '').replace(/\s+/g, '').trim().toUpperCase();
                return txt === 'ALLEMODELLE';
            });
            if (btn) btn.click();
        }''')
        await asyncio.sleep(2.5)

        raw_links = await page.evaluate(r'''() => {
            return Array.from(document.querySelectorAll('a[href*="/marken-modelle/"]'))
                .map(a => a.getAttribute('href'))
                .filter(h => h);
        }''')

        models = []
        seen = set()
        brand_path = brand_url.rstrip('/')
        for href in raw_links:
            full = urljoin(Config.BASE_URL, href).split('#')[0].split('?')[0]
            if full.startswith(brand_path + '/') and full != brand_path + '/':
                segments = full.replace(Config.BASE_URL, '').strip('/').split('/')
                if len(segments) == 3:
                    clean = full.rstrip('/') + '/'
                    if clean not in seen:
                        seen.add(clean)
                        models.append(clean)

        models = list(dict.fromkeys(models))
        if self.test_mode:
            models = models[:2]

        print(f"   Found {len(models)} models")
        return brand_name, models

    async def expand_all_variants(self, page):
        for _ in range(20):
            clicked = await page.evaluate(r'''() => {
                const btns = document.querySelectorAll('.vv__fuelType-dataFooterToggle');
                for (const btn of btns) {
                    if (btn.offsetParent !== null) {
                        btn.click();
                        return true;
                    }
                }
                return false;
            }''')
            if not clicked:
                break
            await asyncio.sleep(1.5)

    async def extract_page_data(self, page):
        data = await page.evaluate(r'''() => {
            const result = {
                headerTitle: 'N/A',
                headerSubtitle: 'N/A',
                fuelTypes: [],
                variants: [],
                editorialTable: null,
                jsonBuildingPeriod: null,
                jsonBodyType: null
            };

            const titleEl = document.querySelector('.vv__header-text-title');
            if (titleEl) result.headerTitle = titleEl.textContent.trim();

            const subtitleEl = document.querySelector('.vv__header-text-subtitle');
            if (subtitleEl) result.headerSubtitle = subtitleEl.textContent.trim();

            const fuelBadges = document.querySelectorAll('.vv__fuelType > .fuelTypeBadge');
            fuelBadges.forEach(badge => {
                const txt = badge.textContent.trim();
                if (txt) result.fuelTypes.push(txt);
            });

            const fuelSections = document.querySelectorAll('.vv__fuelType');
            fuelSections.forEach(section => {
                const badge = section.querySelector('.fuelTypeBadge');
                const fuelType = badge ? badge.textContent.trim() : 'N/A';

                const rows = section.querySelectorAll('.vv__fuelType-dataBodyLine');
                rows.forEach(row => {
                    const cells = row.querySelectorAll('div');
                    if (cells.length >= 5) {
                        const rawVariant = cells[0].textContent.trim();
                        const match = rawVariant.match(/^(.+?)\s+(\d{2}\/\d{4})\s*[–-]\s*(\d{2}\/\d{4})/);
                        const variantName = match ? match[1].trim() : rawVariant;
                        const dateRange = match
                            ? match[2] + ' - ' + match[3]
                            : (rawVariant.match(/seit\s+\d{2}\/\d{4}/) || [''])[0];

                        result.variants.push({
                            fuelType: fuelType,
                            name: variantName,
                            dateRange: dateRange,
                            power: cells[1] ? cells[1].textContent.trim() : 'N/A',
                            acceleration: cells[2] ? cells[2].textContent.trim() : 'N/A',
                            consumption: cells[3] ? cells[3].textContent.trim() : 'N/A',
                            price: cells[4] ? cells[4].textContent.trim() : 'N/A'
                        });
                    }
                });
            });

            const tables = document.querySelectorAll('.editorialTable__table');
            if (tables.length > 0) {
                const table = tables[0];
                const headers = [];
                table.querySelectorAll('.editorialTable__headerCell').forEach(th => {
                    headers.push(th.textContent.trim());
                });
                const rows = [];
                table.querySelectorAll('.editorialTable__bodyRow').forEach(tr => {
                    const label = tr.querySelector('th.firstColumn, th.editorialTable__bodyCell');
                    const vals = [];
                    tr.querySelectorAll('td.editorialTable__bodyCell').forEach(td => {
                        vals.push(td.textContent.trim());
                    });
                    if (label) {
                        rows.push({
                            label: label.textContent.trim(),
                            values: vals
                        });
                    }
                });
                result.editorialTable = { headers: headers, rows: rows };
            }

            try {
                const ctxEl = document.querySelector('#vike_pageContext');
                if (ctxEl) {
                    const ctx = JSON.parse(ctxEl.textContent);
                    const mg = ctx.irContent && ctx.irContent.modelGeneration;
                    if (mg) {
                        if (mg.buildingPeriod) {
                            const from = mg.buildingPeriod.fromYear || '';
                            const till = mg.buildingPeriod.tillYear || '';
                            result.jsonBuildingPeriod = till
                                ? from + ' - ' + till
                                : 'seit ' + from;
                        }
                        if (mg.constructionTypeImages && mg.constructionTypeImages[0]) {
                            result.jsonBodyType = mg.constructionTypeImages[0].type || null;
                        }
                    }
                }
            } catch(e) {}

            return result;
        }''')
        return data

    async def try_extract_hsn_tsn_from_overlay(self, page, variant_index):
        self.api_hsn_tsn = None
        self.api_captured = False

        try:
            clicked = await page.evaluate(r'''(idx) => {
                const rows = document.querySelectorAll('.vv__fuelType-dataBodyLine');
                if (idx < rows.length) {
                    const link = rows[idx].querySelector('.vv__fuelType-dataBodyLineLink');
                    if (link) {
                        link.click();
                        return true;
                    }
                }
                return false;
            }''', variant_index)

            if not clicked: return "N/A"
            await asyncio.sleep(Config.API_WAIT_TIMEOUT)

            result = "N/A"
            if self.api_hsn_tsn:
                hsn, tsn = self._find_hsn_tsn_pair(self.api_hsn_tsn)
                if hsn and tsn: result = f"{hsn}/{tsn}"
                elif hsn and not tsn: result = hsn

            if result == "N/A":
                hsn_from_dom = await page.evaluate(r'''() => {
                    const overlay = document.querySelector('.vvp');
                    if (!overlay) return null;
                    const txt = overlay.innerText || '';
                    const m = txt.match(/HSN\/TSN\s*Schlüsselnummern[\s\S]{0,30}?(\d{4})\s*[\/\-]\s*([A-Z0-9]{2,6})/i);
                    if (m) return m[1] + '/' + m[2];
                    const hsnL = txt.match(/HSN[:\s]*(\d{4})/i);
                    const tsnL = txt.match(/TSN[:\s]*([A-Z0-9]{2,6})/i);
                    if (hsnL && tsnL) return hsnL[1] + '/' + tsnL[1];
                    const gen = txt.match(/(\d{4})\s*[\/\-]\s*([A-Z0-9]{2,6})/);
                    return gen ? gen[1] + '/' + gen[2] : null;
                }''')
                if hsn_from_dom: result = hsn_from_dom

            await page.evaluate("const btn = document.querySelector('.sectionOverlay__buttonClose'); if (btn) btn.click();")
            await asyncio.sleep(0.5)
            return result

        except Exception:
            try:
                await page.evaluate("const btn = document.querySelector('.sectionOverlay__buttonClose'); if (btn) btn.click();")
                await asyncio.sleep(0.5)
            except Exception: pass
        return "N/A"

    def _find_hsn_tsn_pair(self, data, depth=0):
        if depth > 10: return None, None
        if isinstance(data, dict):
            tcert_key = next((k for k in data if 'tcert' in str(k).lower()), None)
            if tcert_key:
                tcert_list = data[tcert_key]
                pairs = []
                if isinstance(tcert_list, list):
                    for entry in tcert_list:
                        if isinstance(entry, dict) and 'Num' in entry and 'Num2' in entry:
                            hv = str(entry.get('Num', '') or '').strip()
                            tv = str(entry.get('Num2', '') or '').strip()
                            if hv and tv and hv != '-' and tv != '-':
                                hsn_clean = re.sub(r'\D', '', hv)[:4]
                                tsn_clean = re.sub(r'[^A-Z0-9]', '', tv.upper())[:6]
                                if hsn_clean and tsn_clean:
                                    pairs.append(f"{hsn_clean}/{tsn_clean}")
                    if pairs:
                        return ','.join(pairs), None
            hsn_k = next((k for k in data if 'hsn' in str(k).lower()), None)
            tsn_k = next((k for k in data if 'tsn' in str(k).lower()), None)
            if hsn_k and tsn_k:
                hv = str(data[hsn_k]).strip() if data[hsn_k] else ''
                tv = str(data[tsn_k]).strip() if data[tsn_k] else ''
                if hv and tv and hv != '-' and tv != '-':
                    hsn_clean = re.sub(r'\D', '', hv)[:4]
                    tsn_clean = re.sub(r'[^A-Z0-9]', '', tv.upper())[:6]
                    if hsn_clean and tsn_clean: return hsn_clean, tsn_clean
            for v in data.values():
                h, t = self._find_hsn_tsn_pair(v, depth + 1)
                if h and t: return h, t
        elif isinstance(data, list):
            for item in data:
                h, t = self._find_hsn_tsn_pair(item, depth + 1)
                if h and t: return h, t
        return None, None

    def build_records(self, brand, model, page_data, fuel_type_filter=None, fallback=False):
        records = []
        body_type = page_data.get('jsonBodyType') or 'N/A'
        body_map = {
            'fliessheck': '掀背車',
            'stufenheck': '斜背車',
            'stlheck_fliessheck': '斜掀背車',
            'steilheck_fliessheck': '高背掀背車',
            'limousine': '轎車',
            'suv': '休旅車',
            'cabrio': '敞篷車',
            'roadster': '敞篷跑車',
            'coupe': '雙門跑車',
            'kombi': '旅行車',
            'kasten': '貨車',
            'van': '廂型車',
            'bus': '巴士',
            'select': '精選型',
            'pritsche': '皮卡',
            'geländewagen': '越野車',
            'gelaendewagen': '越野車',
            'kleinwagen': '小型車'
        }
        if body_type.lower() in body_map:
            body_type = body_map[body_type.lower()]

        fuel_map = {
            'benzin': '汽油',
            'diesel': '柴油',
            'elektro': '電動',
            'elektrischer strom': '電力',
            'benzin/hybrid': '油電混合',
            'benzin/elektro': '油電混合',
            'benzin/elektro-plugin': '插電式油電混合',
            'benzin/elektro-plug': '插電式油電混合',
            'benzin/gas': '汽油/天然氣',
            'benzin/alkohol': '汽油/酒精混合',
            'erdgas': '天然氣',
            'autogas': '液化石油氣',
            'plug-in-hybrid': '插電式油電混合',
            'wasserstoff': '氫燃料'
        }

        tech_rows = {}
        if page_data.get('editorialTable'):
            for row in page_data['editorialTable']['rows']:
                label = row['label'].lower()
                tech_rows[label] = row

        for v in page_data.get('variants', []):
            fuel = v.get('fuelType', 'N/A')
            if fuel_type_filter and fuel != fuel_type_filter:
                continue
            fuel_zh = fuel_map.get(fuel.lower(), fuel)

            v_date = v.get('dateRange', '')
            start_year, end_year = 'N/A', 'N/A'
            if v_date:
                parts = re.split(r'\s*[–-]\s*', v_date.replace('seit ', ''))
                if len(parts) == 2:
                    sy = parts[0].strip().split('/')
                    ey = parts[1].strip().split('/')
                    start_year = sy[1] if len(sy) > 1 else sy[0]
                    end_year = ey[1] if len(ey) > 1 else ey[0]
                elif len(parts) == 1:
                    sy = parts[0].strip().split('/')
                    start_year = sy[1] if len(sy) > 1 else sy[0]
                    end_year = '至今'

            records.append({
                'Brand': brand,
                'Model': model,
                'Category': body_type,
                'Fuel_Type': fuel_zh,
                'Typ': v.get('name', 'N/A'),
                'Start_Year': start_year,
                'End_Year': end_year,
                'HSN_TSN': 'N/A'
            })

        if not records and fallback:
            period = page_data.get('jsonBuildingPeriod') or 'N/A'
            start_year, end_year = 'N/A', 'N/A'
            if period != 'N/A':
                parts = re.split(r'\s*[–-]\s*', period.replace('seit ', ''))
                if len(parts) == 2:
                    start_year, end_year = parts[0].strip(), parts[1].strip()
                elif len(parts) == 1:
                    start_year, end_year = parts[0].strip(), '至今'
            records.append({
                'Brand': brand, 'Model': model, 'Category': body_type,
                'Fuel_Type': 'N/A', 'Typ': f'{model} ({period})' if period != 'N/A' else model,
                'Start_Year': start_year, 'End_Year': end_year, 'HSN_TSN': 'N/A'
            })

        return records

    async def process_model(self, page, brand, model_url):
        model_name = model_url.strip('/').split('/')[-1].replace('-', ' ').title()
        base_brand = brand

        if self.should_stop():
            self.db.flush()
            return False

        await page.goto(model_url, timeout=60000, wait_until="domcontentloaded")
        await dismiss_cookie(page)

        for scroll_y in range(0, 2000, 400):
            await page.evaluate(f"window.scrollTo(0, {scroll_y})")
            await asyncio.sleep(0.3)
        await asyncio.sleep(1.0)

        has_variants = await page.evaluate(r'''() => {
            return document.querySelectorAll('.vv__fuelType-dataBodyLine').length > 0;
        }''')

        if not has_variants:
            page_data = await self.extract_page_data(page)
            records = self.build_records(base_brand, model_name, page_data)
            if not records:
                records = self.build_records(base_brand, model_name, page_data, fallback=True)
            if records:
                saved = self.db.get_progress(base_brand, model_name)
                if saved is not None and saved == len(records) and not self.test_mode:
                    print(f"  [Skip] [{model_name}] (known {saved} records)")
                    return True

                for r in records:
                    self.db.add_to_batch(r)
                self.db.flush()
                if not self.should_stop():
                    self.db.update_progress(base_brand, model_name, len(records))
                print(f"  [OK] [{model_name}] Single spec: {len(records)} records")
            return True

        await self.expand_all_variants(page)

        page_data = await self.extract_page_data(page)

        variant_count = len(page_data.get('variants', []))
        saved_count = self.db.get_progress(base_brand, model_name)

        if saved_count is not None and saved_count == variant_count and not self.test_mode:
            print(f"  [Skip] [{model_name}] (known {saved_count} variants, currently {variant_count})")
            return True

        print(f"  -> [{base_brand} - {model_name}] {variant_count} variants, extracting...")

        records = self.build_records(base_brand, model_name, page_data)

        if self.test_mode:
            records = records[:6]

        for i, record in enumerate(records):
            if self.should_stop():
                break

            sys.stdout.write(
                f"\r      [{i+1}/{len(records)}] {record['Fuel_Type'][:8]} - "
                f"{record['Typ'][:25]}..."
            )
            sys.stdout.flush()

            if record['HSN_TSN'] == 'N/A':
                hsn = await self.try_extract_hsn_tsn_from_overlay(page, i)
                if hsn != 'N/A':
                    record['HSN_TSN'] = hsn

            self.db.add_to_batch(record)

        print()
        
        if not self.should_stop():
            self.db.flush()
            self.db.update_progress(base_brand, model_name, variant_count)
        else:
            self.db.flush()
            return False
            
        return True

    async def run(self):
        print("\n" + "=" * 50)
        print("  AutoBild Scraper v11.0 - Windows Version")
        print(f"  Mode: {'Test' if self.test_mode else 'Full'}")
        if self.target_brand:
            print(f"  Target Brand: {self.target_brand}")
        print(f"  Timeout Protection: {Config.MAX_RUNTIME_HOURS} hours")
        print("=" * 50 + "\n")

        db_bak = Config.DB_FILE.replace('.db', '_backup.db')
        if os.path.exists(Config.DB_FILE):
            try:
                import shutil
                shutil.copy2(Config.DB_FILE, db_bak)
                print(f"  [Backup] DB backup: {db_bak}")
            except Exception as e:
                print(f"  [Backup] DB backup failed: {e}")

        old_snapshot = SnapshotManager.load_snapshot()

        try:
            async with async_playwright() as p:
                browser = await p.chromium.launch(headless=True)
                context = await browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
                    viewport={'width': 1440, 'height': 900}
                )
                context.on("response", self.handle_api_response)
                page = await context.new_page()

                try:
                    brand_urls = await self.collect_brand_urls(page)

                    total_brands = len(brand_urls)
                    for b_idx, b_url in enumerate(brand_urls):
                        if self.should_stop():
                            break

                        brand_name = b_url.strip('/').split('/')[-1].upper()

                        if self.db.is_brand_done(brand_name) and not self.test_mode:
                            print(f"\n  [Skip] [{brand_name}] completed, skipping")
                            continue

                        brand_name, model_urls = await self.collect_model_urls(page, b_url)

                        parsed_ok = True
                        for m_idx, m_url in enumerate(model_urls):
                            if self.should_stop():
                                break

                            try:
                                ok = await self.process_model(page, brand_name, m_url)
                                if not ok: parsed_ok = False
                            except Exception as e:
                                print(f"    [Error] Model processing failed: {e}")
                                parsed_ok = False

                        if not self.should_stop():
                            csv_count = self.db.export_brand_csv(brand_name)
                            elapsed = (time.time() - self.start_time) / 60
                            print(f"\n   [Export] {brand_name} done ({csv_count} records) "
                                  f"[{b_idx+1}/{total_brands} brands] [{elapsed:.1f} min]")
                            if parsed_ok:
                                self.db.mark_brand_done(brand_name)

                except Exception as e:
                    print(f"\n[Error] Execution interrupted: {e}")
                finally:
                    self.db.flush()
                    await browser.close()

        except Exception as e:
            print(f"\n[Error] Playwright failed to start: {e}")
            print("\nThis is a Windows/Python compatibility issue.")
            print("Please try these solutions:")
            print("\nSolution 1 - Reinstall Playwright:")
            print("  pip uninstall playwright -y")
            print("  pip install playwright==1.40.0")
            print("  python -m playwright install chromium")
            print("\nSolution 2 - Use newer Python:")
            print("  Download Python 3.10+ from python.org")
            print("\nSolution 3 - Run on Mac/Linux instead")
            sys.exit(1)

        new_snapshot = self.db.get_all_data()
        if self.diff and not new_snapshot.empty:
            SnapshotManager.generate_diff_report(old_snapshot, new_snapshot)

        self.db.close()

        elapsed = (time.time() - self.start_time) / 60
        print(f"\n[Complete] Scraper finished! Time: {elapsed:.1f} minutes")
        print(f"[Database] {Config.DB_FILE}")
        print(f"[CSV Dir]  {Config.CSV_DIR}/")


def show_status():
    if not os.path.exists(Config.DB_FILE):
        print("[Error] Database not found, please run scraper first")
        return
    conn = sqlite3.connect(Config.DB_FILE)
    try:
        df = pd.read_sql_query('''
            SELECT Brand, Fuel_Type, Category,
                   COUNT(DISTINCT Model) as Models,
                   COUNT(*) as Rows
            FROM view_car_catalog
            GROUP BY Brand, Fuel_Type, Category
            ORDER BY Brand, Fuel_Type
        ''', conn)
        if df.empty:
            print("[Info] Database is empty")
        else:
            print("\n[Statistics]")
            print(df.to_string())
            total = pd.read_sql_query("SELECT COUNT(*) as Total FROM view_car_catalog", conn)
            print(f"\n  Total Records: {total.iloc[0]['Total']}")
    except Exception as e:
        print(f"[Error] Read failed: {e}")
    finally:
        conn.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="AutoBild Scraper v11.0 - Windows")
    parser.add_argument("--test", action="store_true", help="Test mode")
    parser.add_argument("--reset", action="store_true", help="Reset database and snapshot")
    parser.add_argument("--status", action="store_true", help="Show statistics")
    parser.add_argument("--brand", type=str, default=None, help="Scrape specific brand only")
    parser.add_argument("--diff", action="store_true", default=True, help="Generate diff report (default: on)")
    parser.add_argument("--no-diff", action="store_false", dest="diff", help="Skip diff report")
    args = parser.parse_args()

    if args.reset:
        for f in [Config.DB_FILE]:
            if os.path.exists(f):
                os.remove(f)
        for f in glob.glob(os.path.join(Config.CSV_DIR, 'autobild_snapshot_*.xlsx')):
            os.remove(f)
        for pat in ['autobild_snapshot*.xlsx', 'autobild_diff*.xlsx']:
            for f in glob.glob(os.path.join(Config.CSV_DIR, pat)):
                os.remove(f)
        print("[OK] Database and snapshot reset!")
        sys.exit(0)

    if args.status:
        show_status()
        sys.exit(0)

    scraper = AutoBildScraper(test_mode=args.test, target_brand=args.brand, diff=args.diff)
    asyncio.run(scraper.run())
