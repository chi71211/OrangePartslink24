import sys
import os

# ==========================================
# 解決 Windows 終端機中文亂碼問題
# ==========================================
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter("utf-8")(sys.stdout.detach())
    sys.stderr = codecs.getwriter("utf-8")(sys.stderr.detach())
else:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import sqlite3
import requests
import pandas as pd
import re
import json
import time
import random
import shutil
from datetime import datetime
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from tqdm import tqdm
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ==========================================
# 全域設定 (檔案路徑與變數 V10)
# ==========================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(SCRIPT_DIR, '胎壓檢測器資料庫_V10.db')
SQL_PATH = os.path.join(SCRIPT_DIR, '胎壓檢測器資料庫_V10.sql')
PROGRESS_FILE = os.path.join(SCRIPT_DIR, 'scrape_progress_V10.json')

session = requests.Session()
retry_strategy = Retry(total=5, backoff_factor=1.5, status_forcelist=[429, 500, 502, 503, 504])
adapter = HTTPAdapter(max_retries=retry_strategy)
session.mount("https://", adapter)
session.mount("http://", adapter)
session.headers.update({
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "de",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Connection": "keep-alive"
})

GLOBAL_OLD_DB_DF = pd.DataFrame()

# ==========================================
# 1. 進度與週期管理
# ==========================================
def load_progress():
    if os.path.exists(PROGRESS_FILE):
        try:
            with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return {"cycle_start": time.time(), "last_brand": "", "last_class": "", "last_tg": ""}

def save_progress(brand="", car_class="", tg_id="", completed=False):
    prog = load_progress()
    if completed:
        prog["cycle_start"] = time.time()
        prog.update({"last_brand": "", "last_class": "", "last_tg": ""})
    else:
        prog["last_brand"] = brand
        prog["last_class"] = car_class
        prog["last_tg"] = tg_id
    with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
        json.dump(prog, f)

def check_30_day_cycle():
    prog = load_progress()
    if time.time() - prog.get("cycle_start", 0) > 2592000:
        print("\n⏳ [週期檢查] 超過 30 天，啟動全面掃描（保留歷史資料）...")
        fresh = {"cycle_start": time.time(), "last_brand": "", "last_class": "", "last_tg": ""}
        with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
            json.dump(fresh, f)
        return fresh, True
    return prog, False

# ==========================================
# 2. 安全網路請求 & API 輔助函式
# ==========================================
def safe_json_get(url, params=None, timeout=15):
    try:
        r = session.get(url, params=params, timeout=timeout)
        r.raise_for_status()
        return r.json() if r.text and r.text.strip() else []
    except Exception as e:
        if "429" in str(e):
            print("⚠️ 觸發網站的 Rate Limit (429) 保護機制，強制等待 10 秒...")
            time.sleep(10)
        else:
            print(f"請求失敗 {url}: {e}")
        time.sleep(1.5)
        return []

def get_manufacturers(): return safe_json_get("https://www.interpneu-raederkonfigurator.de/api/cars/manufacturers")
def get_classes(brand): return safe_json_get("https://www.interpneu-raederkonfigurator.de/api/cars/classes", {"manufacturer": brand})
def get_type_groups(brand, car_class): 
    return safe_json_get("https://www.interpneu-raederkonfigurator.de/api/cars/type-groups", {"manufacturer": brand, "class": car_class})
def get_versions(tg): 
    return safe_json_get("https://www.interpneu-raederkonfigurator.de/api/cars/version-groups", {"group": tg})
def get_car_hsn_tsn(tag): return safe_json_get("https://www.interpneu-raederkonfigurator.de/api/cars/car", {"carTag": tag})
def get_tpms(tag): return safe_json_get("https://www.interpneu-raederkonfigurator.de/api/tpms/carTpms", {"carTag": tag})

def get_sensor_details(manufacturer_ids):
    if not manufacturer_ids: return {}
    ids_str = ",".join(map(str, manufacturer_ids)) if isinstance(manufacturer_ids, list) else str(manufacturer_ids)
    return safe_json_get("https://www.interpneu-raederkonfigurator.de/api/gpsr/data", {"manufacturerIds": ids_str})

def format_year(d): 
    return "至今" if not d or d == "0000-00-00" else d[:7]

def sanitize_filename(name):
    name = re.sub(r'[\\/*?:"<>|]', '_', name)
    return name.replace('ü','ue').replace('ä','ae').replace('ö','oe').replace('ß','ss')

def find_key_value(d, keywords):
    if isinstance(d, dict):
        for k, v in d.items():
            if any(kw in str(k).lower() for kw in keywords) and not isinstance(v, (dict, list)):
                if v and str(v).strip(): return str(v).strip()
            res = find_key_value(v, keywords)
            if res: return res
    elif isinstance(d, list):
        for item in d:
            res = find_key_value(item, keywords)
            if res: return res
    return ""

def add_space_after_comma(text):
    if not isinstance(text, str): return text
    return ', '.join([s.strip() for s in text.split(',') if s.strip()])

# ==========================================
# 3. 資料庫升級與寫入邏輯
# ==========================================
def upgrade_db_schema():
    if not os.path.exists(DB_PATH): return
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("PRAGMA table_info(tpms_sensors)")
        columns = [info[1] for info in cursor.fetchall()]
        if len(columns) > 0 and "OE Nummer" not in columns:
            cursor.execute('ALTER TABLE tpms_sensors ADD COLUMN "OE Nummer" TEXT DEFAULT ""')
            print("🔧 [資料庫升級] 自動為舊資料庫擴充 'OE Nummer' 欄位，保留原有資料。")
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"⚠️ 資料庫升級檢查失敗: {e}")

def save_batch_to_sql(batch_data):
    if not batch_data: return
    df = pd.DataFrame(batch_data).fillna("")

    group_cols = ['Brand', 'Model', 'Typ', 'Start Year', 'End Year', 'HSN', 'TSN', 'OE sensor', 'created date', 'Manufacturer', 'Frequency', 'OE Nummer']
    df_merged = df.drop_duplicates(subset=group_cols)
    
    cols = ['Brand', 'Model', 'Typ', 'Start Year', 'End Year', 'HSN', 'TSN',
            'created date', 'OE sensor', 'OE Nummer', 'Manufacturer', 'Frequency']
    df_merged = df_merged.reindex(columns=cols)

    try:
        conn = sqlite3.connect(DB_PATH)
        conn.execute('''CREATE TABLE IF NOT EXISTS tpms_sensors (
            "Brand" TEXT, "Model" TEXT, "Typ" TEXT, "Start Year" TEXT, "End Year" TEXT,
            "HSN" TEXT, "TSN" TEXT, "created date" TEXT,
            "OE sensor" TEXT, "OE Nummer" TEXT, "Manufacturer" TEXT, "Frequency" TEXT,
            UNIQUE("Brand", "Model", "Typ", "Start Year", "End Year", "HSN", "TSN", "OE sensor", "created date")
        )''')
        conn.executemany('''
            REPLACE INTO tpms_sensors
            ("Brand", "Model", "Typ", "Start Year", "End Year", "HSN", "TSN", "created date",
             "OE sensor", "OE Nummer", "Manufacturer", "Frequency")
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', df_merged.values.tolist())
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"⚠️ DB 寫入失敗: {e}")

def auto_export_sql():
    if not os.path.exists(DB_PATH): return
    conn = sqlite3.connect(DB_PATH)
    with open(SQL_PATH, 'w', encoding='utf-8') as f:
        for line in conn.iterdump():
            f.write(f'{line}\n')
    conn.close()
    print(f"🎉 SQL 備份完成: {os.path.basename(SQL_PATH)}")

# ==========================================
# 4. 更新差異報表產出 (Diff) - O(舊)與N(新)並排模式
# ==========================================
def _new_records_from_df(brand, df):
    compare_cols = ['HSN', 'TSN', 'created date', 'OE sensor', 'OE Nummer', 'Manufacturer', 'Frequency']
    records = []
    for _, row in df.iterrows():
        rec = {
            '變更狀態': '新增 (New)', 'Brand': row['Brand'], 'Model': row['Model'],
            'Typ': row['Typ'], '年份': row['年份']
        }
        for col in compare_cols:
            rec[f'{col} (O)'] = ""
            rec[f'{col} (N)'] = str(row[col]) if col in df.columns and pd.notna(row[col]) else ""
        records.append(rec)
    return records

def export_diff_excel(brand, old_df, new_df):
    keys = ['Brand', 'Model', 'Typ', '年份']
    compare_cols = ['HSN', 'TSN', 'created date', 'OE sensor', 'OE Nummer', 'Manufacturer', 'Frequency']

    if old_df.empty:
        return _new_records_from_df(brand, new_df)

    old_b = old_df[old_df['Brand'] == brand].copy()
    if old_b.empty:
        return _new_records_from_df(brand, new_df)

    old_b = old_b.set_index(keys)
    new_b = new_df.set_index(keys)

    new_idx = new_b.index.difference(old_b.index)
    common_idx = new_b.index.intersection(old_b.index)

    diff_records = []

    for idx in common_idx:
        old_row = old_b.loc[idx]
        new_row = new_b.loc[idx]
        if isinstance(old_row, pd.DataFrame) or isinstance(new_row, pd.DataFrame):
            continue
        is_changed = False
        record = {'變更狀態': '修改 (Updated)', 'Brand': idx[0], 'Model': idx[1], 'Typ': idx[2], '年份': idx[3]}
        for col in compare_cols:
            val_o = str(old_row[col]) if col in old_b.columns and pd.notna(old_row[col]) else ""
            val_n = str(new_row[col]) if col in new_b.columns and pd.notna(new_row[col]) else ""
            record[f'{col} (O)'] = val_o
            record[f'{col} (N)'] = val_n
            if val_o != val_n: is_changed = True
        if is_changed:
            diff_records.append(record)

    for idx in new_idx:
        new_row = new_b.loc[idx]
        if isinstance(new_row, pd.DataFrame):
            continue
        record = {'變更狀態': '新增 (New)', 'Brand': idx[0], 'Model': idx[1], 'Typ': idx[2], '年份': idx[3]}
        for col in compare_cols:
            record[f'{col} (O)'] = ""
            record[f'{col} (N)'] = str(new_row[col]) if col in new_b.columns and pd.notna(new_row[col]) else ""
        diff_records.append(record)

    if not diff_records:
        return []

    tqdm.write(f"📝 發現 {len(diff_records)} 筆變更 ({brand})")
    return diff_records

# ==========================================
# 5. 變更紀錄.xlsx — 每次新增一個分頁（日期命名）
# ==========================================
def _merge_diff_records(existing, new_records):
    seen = {}
    for r in existing + new_records:
        key = (str(r.get('變更狀態','')), str(r.get('Brand','')), str(r.get('Model','')), str(r.get('Typ','')))
        seen[key] = r
    return list(seen.values())

def append_diff_to_excel(output_dir, all_diff_records):
    today_str = datetime.now().strftime('%Y%m%d')

    output_path = os.path.join(output_dir, "變更紀錄.xlsx")
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if today_str in wb.sheetnames:
        existing = []
        ws_old = wb[today_str]
        old_headers = [ws_old.cell(1, c).value for c in range(1, ws_old.max_column+1)]
        for r in range(2, ws_old.max_row+1):
            row_data = {}
            for c, h in enumerate(old_headers, 1):
                row_data[h] = ws_old.cell(r, c).value
            if row_data:
                existing.append(row_data)
        del wb[today_str]
        all_diff_records = _merge_diff_records(existing, all_diff_records)

    ws = wb.create_sheet(title=today_str)

    if not all_diff_records:
        ws.cell(row=1, column=1, value="本日無任何變更")
        wb.save(output_path)
        print(f"📊 變更紀錄已新增分頁 [{today_str}] → 無變更")
        return

    diff_df = pd.DataFrame(all_diff_records)
    diff_df = diff_df.drop_duplicates(subset=['變更狀態', 'Brand', 'Model', 'Typ'])

    compare_cols = ['HSN', 'TSN', 'created date', 'OE sensor', 'OE Nummer', 'Manufacturer', 'Frequency']
    cols_order = ['變更狀態', 'Brand', 'Model', 'Typ', '年份']
    for col in compare_cols:
        cols_order.append(f'{col} (O)')
        cols_order.append(f'{col} (N)')
    diff_df = diff_df[cols_order]

    red_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
    red_font = Font(color='D32F2F', bold=True)
    green_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    green_font = Font(color='2E7D32', bold=True)
    orange_font = Font(color='E65100', bold=True)
    header_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    header_font = Font(bold=True)
    new_row_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    new_row_font = Font(color='1B5E20')

    headers = list(diff_df.columns)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for r_idx, (_, row) in enumerate(diff_df.iterrows(), 2):
        status = row['變更狀態']
        is_new = (status == '新增 (New)')
        for col_idx, col_name in enumerate(headers, 1):
            val = row[col_name]
            if val is None or (isinstance(val, float) and pd.isna(val)):
                val = ""
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            if is_new:
                cell.fill = new_row_fill
                cell.font = new_row_font
            elif col_name == '變更狀態':
                cell.font = orange_font
            elif col_name.endswith(' (O)') or col_name.endswith(' (N)'):
                base = col_name[:-4]
                col_o = f'{base} (O)'
                col_n = f'{base} (N)'
                if col_o in headers and col_n in headers:
                    o_val = row[col_o]
                    n_val = row[col_n]
                    if str(o_val) != str(n_val):
                        if col_name.endswith(' (O)'):
                            cell.fill = red_fill
                            cell.font = red_font
                        else:
                            cell.fill = green_fill
                            cell.font = green_font

    wb.save(output_path)
    print(f"📊 變更紀錄已新增分頁 [{today_str}] → {os.path.basename(output_path)}")

# ==========================================
# 6. DB 備份 (每次執行開始前備份目前 DB)
# ==========================================
def backup_db():
    if not os.path.exists(DB_PATH):
        print("📦 DB 不存在，略過備份")
        return
    backup_path = os.path.join(SCRIPT_DIR, "胎壓檢測器資料庫_V10_備份.db")
    try:
        shutil.copy2(DB_PATH, backup_path)
        print(f"📦 DB 備份完成 → {os.path.basename(backup_path)}")
    except Exception as e:
        print(f"⚠️ DB 備份失敗: {e}")

def _brand_df_to_excel(brand_df, brand, folder_path):
    for col in ['HSN','TSN','created date','OE sensor','OE Nummer','Manufacturer','Frequency']:
        if col in brand_df.columns:
            brand_df[col] = brand_df[col].apply(add_space_after_comma)
    if '年份' not in brand_df.columns:
        brand_df['年份'] = brand_df['Start Year'].astype(str) + " ~ " + brand_df['End Year'].astype(str)
    order = ['Brand','Model','Typ','年份','HSN','TSN','created date','OE sensor','OE Nummer','Manufacturer','Frequency']
    brand_df = brand_df.reindex(columns=order)
    path = os.path.join(folder_path, f"{sanitize_filename(brand)}_Data.xlsx")
    brand_df.to_excel(path, index=False)

def _ensure_brand_excel(brand, folder_path):
    brand_path = os.path.join(folder_path, f"{sanitize_filename(brand)}_Data.xlsx")
    if os.path.exists(brand_path):
        return
    try:
        conn = sqlite3.connect(DB_PATH)
        bdf = pd.read_sql_query('SELECT * FROM view_tpms_sensors_unique1 WHERE "Brand"=?', conn, params=(brand,))
        conn.close()
        if not bdf.empty:
            _brand_df_to_excel(bdf, brand, folder_path)
            print(f"📄 補產 Excel: {brand} → {len(bdf)} 筆")
    except Exception as e:
        print(f"⚠️ 補產 Excel 失敗 {brand}: {e}")

def _write_brand_diff(brand, folder_path, all_diff_records):
    try:
        conn = sqlite3.connect(DB_PATH)
        df = pd.read_sql_query('SELECT * FROM view_tpms_sensors_unique1 WHERE "Brand"=?', conn, params=(brand,))
        conn.close()
        if not df.empty:
            for col in ['HSN', 'TSN', 'created date', 'OE sensor', 'OE Nummer', 'Manufacturer', 'Frequency']:
                if col in df.columns:
                    df[col] = df[col].apply(add_space_after_comma)
            df['年份'] = df['Start Year'].astype(str) + " ~ " + df['End Year'].astype(str)
            _brand_df_to_excel(df.copy(), brand, folder_path)
        new_recs = export_diff_excel(brand, GLOBAL_OLD_DB_DF, df)
        all_diff_records.extend(new_recs)
        append_diff_to_excel(folder_path, all_diff_records)
    except Exception as e:
        print(f"⚠️ 寫入品牌資料失敗 {brand}: {e}")

def recover_missing_excels(folder_path):
    if not os.path.exists(DB_PATH):
        return
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.execute('SELECT DISTINCT "Brand" FROM view_tpms_sensors_unique1')
        db_brands = [row[0] for row in cursor.fetchall()]
        conn.close()
        for b in db_brands:
            _ensure_brand_excel(b, folder_path)
        print("✅ 復原檢查完成")
    except Exception as e:
        print(f"⚠️ 復原檢查異常: {e}")

# ==========================================
# 7. 主程式 (爬蟲核心迴圈)
# ==========================================
def main_scraper_all():
    global GLOBAL_OLD_DB_DF
    folder_path = os.path.join(SCRIPT_DIR, sanitize_filename("胎壓檢測器資料庫_V10"))
    os.makedirs(folder_path, exist_ok=True)
    
    upgrade_db_schema()

    backup_db()

    if os.path.exists(DB_PATH):
        try:
            conn = sqlite3.connect(DB_PATH)
            GLOBAL_OLD_DB_DF = pd.read_sql_query('SELECT * FROM view_tpms_sensors_unique1', conn)
            conn.close()
            
            if not GLOBAL_OLD_DB_DF.empty:
                for col in ['HSN', 'TSN', 'created date', 'OE sensor', 'OE Nummer', 'Manufacturer', 'Frequency']:
                    if col in GLOBAL_OLD_DB_DF.columns:
                        GLOBAL_OLD_DB_DF[col] = GLOBAL_OLD_DB_DF[col].apply(add_space_after_comma)
                if 'Start Year' in GLOBAL_OLD_DB_DF.columns and 'End Year' in GLOBAL_OLD_DB_DF.columns:
                    GLOBAL_OLD_DB_DF['年份'] = GLOBAL_OLD_DB_DF['Start Year'].astype(str) + " ~ " + GLOBAL_OLD_DB_DF['End Year'].astype(str)
            
            recover_missing_excels(folder_path)
        except Exception as e:
            print(f"⚠️ 載入舊資料庫比對檔發生錯誤: {e}")

    brands = get_manufacturers()
    if not isinstance(brands, list) or not brands:
        print("❌ 無法取得Brand清單")
        return

    prog, _ = check_30_day_cycle()
    skip_mode = bool(prog.get("last_brand"))

    conn = sqlite3.connect(DB_PATH)
    conn.execute('''CREATE TABLE IF NOT EXISTS tpms_sensors (
        "Brand" TEXT, "Model" TEXT, "Typ" TEXT, "Start Year" TEXT, "End Year" TEXT,
        "HSN" TEXT, "TSN" TEXT, "created date" TEXT, 
        "OE sensor" TEXT, "OE Nummer" TEXT, "Manufacturer" TEXT, "Frequency" TEXT,
        UNIQUE("Brand", "Model", "Typ", "Start Year", "End Year", "HSN", "TSN", "OE sensor", "created date")
    )''')
    
    conn.execute('DROP VIEW IF EXISTS view_tpms_sensors_unique1;')
    conn.execute('''
        CREATE VIEW view_tpms_sensors_unique1 AS
        SELECT
            "Brand",
            "Model",
            "Typ",
            "Start Year",
            "End Year",
            GROUP_CONCAT(DISTINCT "HSN") AS "HSN",
            GROUP_CONCAT(DISTINCT "TSN") AS "TSN",
            GROUP_CONCAT(DISTINCT "created date") AS "created date",
            GROUP_CONCAT(DISTINCT "OE sensor") AS "OE sensor",
            GROUP_CONCAT(DISTINCT "OE Nummer") AS "OE Nummer",
            GROUP_CONCAT(DISTINCT "Manufacturer") AS "Manufacturer",
            GROUP_CONCAT(DISTINCT "Frequency") AS "Frequency"
        FROM tpms_sensors
        GROUP BY "Brand", "Model", "Typ", "Start Year", "End Year";
    ''')
    conn.commit()
    conn.close()

    batch_data = []
    completed_brands = []
    today_str = datetime.now().strftime('%Y%m%d')
    all_diff_records = []

    print(f"🔍 共 {len(brands)} 個Brand\n")

    try:
        for brand in tqdm(brands, desc="總進度", ncols=100):
            if skip_mode and brand != prog.get("last_brand"):
                tqdm.write(f"⏭️ 已完成: {brand} ✅")
                completed_brands.append(brand)
                _ensure_brand_excel(brand, folder_path)
                continue

            if skip_mode and brand == prog.get("last_brand"):
                skip_mode = False
                tqdm.write(f"▶️ 從 {brand} 繼續...")

            tqdm.write(f"\n🚗 正在處理: 【{brand}】")
            save_progress(brand=brand)

            classes = get_classes(brand)
            if not isinstance(classes, list): classes = []

            for car_class in tqdm(classes, desc=f"{brand} Model", leave=False, ncols=80):
                type_groups = get_type_groups(brand, car_class)
                if not isinstance(type_groups, list): type_groups = []

                for tg_data in type_groups:
                    tg_id = tg_data.get("group") if isinstance(tg_data, dict) else None
                    if not tg_id: continue

                    if skip_mode and tg_id != prog.get("last_tg"): continue
                    if skip_mode: skip_mode = False

                    save_progress(brand, car_class, tg_id)

                    versions = get_versions(tg_id)
                    if not isinstance(versions, list): continue
                    versions.sort(key=lambda x: x.get('productionFrom', ''), reverse=True)

                    for version in versions:
                        car_tag = str(version.get("tag") or version.get("carTag") or "")
                        if not car_tag: continue

                        year_from = format_year(version.get("productionFrom"))
                        year_to = format_year(version.get("productionTo"))
                        model_version = version.get("version", "")

                        time.sleep(random.uniform(0.65, 1.35))

                        try:
                            car_details = get_car_hsn_tsn(car_tag)
                            hsn = car_details.get("hsn", "") if isinstance(car_details, dict) else ""
                            tsn = car_details.get("tsn", "") if isinstance(car_details, dict) else ""

                            tpms_data = get_tpms(car_tag)
                            oe_list = []

                            if isinstance(tpms_data, dict) and "tpms" in tpms_data:
                                sensor_ids = []
                                for s in tpms_data["tpms"]:
                                    if s.get("oeAm") == "O":
                                        m_id = s.get("manufacturerId") or s.get("articleId") or s.get("id") or s.get("tpmsId")
                                        if m_id: sensor_ids.append(m_id)
                                
                                details_dict = {}
                                if sensor_ids:
                                    time.sleep(random.uniform(0.3, 0.7)) 
                                    detailed_info = get_sensor_details(sensor_ids)
                                    if isinstance(detailed_info, list):
                                        for item in detailed_info:
                                            item_id = item.get("manufacturerId") or item.get("id") or item.get("articleId")
                                            if item_id: details_dict[str(item_id)] = item
                                    elif isinstance(detailed_info, dict):
                                        data_list = detailed_info.get("data", []) or detailed_info.get("items", [])
                                        if isinstance(data_list, list):
                                            for item in data_list:
                                                item_id = item.get("manufacturerId") or item.get("id") or item.get("articleId")
                                                if item_id: details_dict[str(item_id)] = item
                                        else:
                                            details_dict = detailed_info

                                for s in tpms_data["tpms"]:
                                    if s.get("oeAm") == "O":
                                        m_id = str(s.get("manufacturerId") or s.get("articleId") or s.get("id") or s.get("tpmsId") or "")
                                        s_detail = details_dict.get(m_id, {}) if m_id else {}

                                        hersteller = s.get("manufacturer", "") or s_detail.get("manufacturerName1", "")
                                        frequenz = str(s.get("tpmsFrequency", "") or s_detail.get("frequenz", ""))
                                        oe_nummer = s.get("oeNumber", "") or s.get("equipment", "")
                                        
                                        if not frequenz:
                                            sd = str(s).lower() + str(s_detail).lower()
                                            frequenz = '433' if '433' in sd else '434' if '434' in sd else '315' if '315' in sd else ''
                                        
                                        baujahr = s_detail.get("baujahr") or s.get("baujahr", "")
                                        if not baujahr:
                                            s_str = str(s) + str(s_detail)
                                            match = re.search(r'(\d{2}/\d{4}\s*-(\s*\d{2}/\d{4})?)', s_str)
                                            if match:
                                                baujahr = match.group(1).strip()
                                            else:
                                                baujahr = find_key_value(s, ['baujahr'])
                                                
                                        if not baujahr or baujahr == "0000-00-00":
                                            baujahr = f"{year_from} ~ {year_to}"
                                            
                                        oe_list.append({
                                            "oe": str(s.get("tpmsDescFrontend", "")),
                                            "oe_nummer": str(oe_nummer),
                                            "baujahr": baujahr,
                                            "hersteller": str(hersteller),
                                            "frequenz": str(frequenz)
                                        })

                            if not oe_list:
                                oe_list = [{"oe":"", "oe_nummer":"", "baujahr":"", "hersteller":"", "frequenz":""}]

                            for info in oe_list:
                                batch_data.append({
                                    "Brand": brand, "Model": car_class, "Typ": model_version,
                                    "Start Year": year_from, "End Year": year_to,
                                    "HSN": hsn, "TSN": tsn,
                                    "created date": info["baujahr"],
                                    "OE sensor": info["oe"],
                                    "OE Nummer": info["oe_nummer"],
                                    "Manufacturer": info["hersteller"],
                                    "Frequency": info["frequenz"]
                                })

                            if len(batch_data) >= 80:
                                save_batch_to_sql(batch_data)
                                batch_data.clear()

                        except Exception as e:
                            tqdm.write(f"⚠️ 異常 {model_version}: {e}")

                if batch_data:
                    save_batch_to_sql(batch_data)
                    batch_data.clear()

                _write_brand_diff(brand, folder_path, all_diff_records)
                tqdm.write(f"   ✅ Model完成: {car_class}")

            completed_brands.append(brand)
            tqdm.write(f"🎉 Brand完成: 【{brand}】 ✅")

    finally:
        if batch_data:
            save_batch_to_sql(batch_data)
            batch_data.clear()
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.execute('SELECT DISTINCT "Brand" FROM view_tpms_sensors_unique1')
            db_brands = [row[0] for row in cursor.fetchall()]
            conn.close()
            for b in db_brands:
                _ensure_brand_excel(b, folder_path)
        except Exception as e:
            print(f"⚠️ 中斷補產異常: {e}")
        append_diff_to_excel(folder_path, all_diff_records)

    print("\n" + "="*60)
    print("🎊 爬蟲任務全部完成！")
    save_progress(completed=True)
    print(f"✅ 此次共完整處理 {len(completed_brands)} 個Brand")
    print("="*60)

    auto_export_sql()

if __name__ == "__main__":
    main_scraper_all()