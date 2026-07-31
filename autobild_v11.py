"""
AutoBild 爬蟲系統 v11.0 - 本地 VS Code 完美中斷儲存版 (含中文翻譯)
========================================
核心修正：
1. 訊號攔截：支援 VS Code 手動按 Ctrl+C 或是關閉終端機時，安全存檔不遺失。
2. 7日與完工規則：7日內接續掃描；若超過7日未打開、或已全部掃描完畢，則下次自動從頭開始。
3. 混合解析架構：精準定位 vv__ / vvp__，並解析 #vike_pageContext 取得底層 JSON。
4. API 攔截器：從動態載入的 API 回應中提取 HSN/TSN。
5. 中文翻譯：自動將車身與燃料類型翻譯為繁體中文。

用法：
  python autobild_v11.py              # 完整掃描 (支援中斷續傳)
  python autobild_v11.py --test       # 測試模式 (2 廠牌 x 2 車系)
  python autobild_v11.py --reset      # 強制重置資料庫
  python autobild_v11.py --status     # 查看目前統計
   python autobild_v11.py --brand VW   # 只抓特定品牌
   python autobild_v11.py --no-diff    # 跳過差異比對
"""

import subprocess
import sys
import os
import glob
import signal

def install_packages():
    packages = ['nest_asyncio', 'playwright', 'pandas', 'openpyxl']
    for pkg in packages:
        try:
            __import__(pkg.replace('-', '_'))
        except ImportError:
            print(f"Installing {pkg}...")
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', pkg, '-q'])
    try:
        import playwright
        subprocess.check_call([sys.executable, '-m', 'playwright', 'install', 'chromium'], 
                            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception:
        pass

install_packages()

import nest_asyncio
nest_asyncio.apply()
import asyncio
import time
import random
import sqlite3
import json
import argparse
import pandas as pd
import re
from datetime import datetime, timedelta
from urllib.parse import urljoin
from playwright.async_api import async_playwright
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font

os.chdir(os.path.dirname(os.path.abspath(__file__)))


class Config:
    BASE_URL = "https://www.autobild.de"
    CATALOG_URL = f"{BASE_URL}/marken-modelle/#aktuell"
    CSV_DIR = "AutoBild_Exports"
    DB_FILE = "autobild_master.db"
    BATCH_SIZE = 20
    DELAY_MIN = 0.8
    DELAY_MAX = 1.6
    MAX_RETRIES = 3
    ENABLE_RUNTIME_LIMIT = False
    MAX_RUNTIME_HOURS = 5.5
    API_WAIT_TIMEOUT = 8
    SNAPSHOT_FILE = os.path.join(CSV_DIR, "autobild_snapshot.xlsx")
    DIFF_FILE = os.path.join(CSV_DIR, "autobild_diff.xlsx")


class DatabaseManager:
    def __init__(self):
        os.makedirs(Config.CSV_DIR, exist_ok=True)
        self.conn = sqlite3.connect(Config.DB_FILE)
        self.cursor = self.conn.cursor()
        self.batch = []
        self._init_db()
        self.check_and_reset_progress()
        self.changelog_path = Config.CSV_DIR + '/autobild_changelog.xlsx'
        self._init_changelog()

    def _init_changelog(self):
        try:
            if os.path.exists(self.changelog_path):
                wb = load_workbook(self.changelog_path)
                if 'Changelog' not in wb.sheetnames:
                    ws = wb.create_sheet('Changelog', 0)
                    ws.append(['Timestamp', 'Action', 'Brand', 'Model', 'Category', 'Fuel_Type', 'Typ', 'Start_Year', 'End_Year', 'HSN_TSN_Old', 'HSN_TSN_New'])
                    ws.column_dimensions['A'].width = 20
                    for col in ['B','C','D','E','F','G','H','I','J','K']:
                        ws.column_dimensions[col].width = 16
                wb.close()
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = 'Changelog'
                ws.append(['Timestamp', 'Action', 'Brand', 'Model', 'Category', 'Fuel_Type', 'Typ', 'Start_Year', 'End_Year', 'HSN_TSN_Old', 'HSN_TSN_New'])
                ws.column_dimensions['A'].width = 20
                for col in ['B','C','D','E','F','G','H','I','J','K']:
                    ws.column_dimensions[col].width = 16
                wb.save(self.changelog_path)
                wb.close()
        except Exception:
            pass

    def _log_changelog(self, records):
        try:
            wb = load_workbook(self.changelog_path)
            ws = wb['Changelog']
            green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            for action, r in records:
                ts = time.strftime('%Y-%m-%d %H:%M:%S')
                row_data = [ts, action, r.get('Brand',''), r.get('Model',''), r.get('Category',''), r.get('Fuel_Type',''), r.get('Typ',''), r.get('Start_Year',''), r.get('End_Year',''), r.get('_hsn_old',''), r.get('HSN_TSN','')]
                ws.append(row_data)
                new_row = ws.max_row
                fill = green if action == 'INSERT' else red
                for col in range(1, len(row_data) + 1):
                    ws.cell(row=new_row, column=col).fill = fill
            wb.save(self.changelog_path)
            wb.close()
        except Exception as e:
            print(f"  [Changelog] 寫入失敗: {e}")

    def _init_db(self):
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS car_catalog (
                Brand TEXT, Model TEXT, Category TEXT, Fuel_Type TEXT, Typ TEXT, 
                Start_Year TEXT, End_Year TEXT, HSN_TSN TEXT,
                UNIQUE(Brand, Model, Category, Fuel_Type, Typ, Start_Year, End_Year, HSN_TSN)
            )
        ''')

        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS model_progress (
                Brand TEXT, Model TEXT, variant_count INTEGER, last_scraped TEXT,
                PRIMARY KEY(Brand, Model)
            )
        ''')

        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS system_metadata (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        ''')

        self.cursor.execute('DROP VIEW IF EXISTS view_car_catalog')
        self.cursor.execute('''
            CREATE VIEW view_car_catalog AS
            SELECT Brand, Model, Category, Fuel_Type, Typ, Start_Year, End_Year, HSN_TSN
            FROM car_catalog
            GROUP BY Brand, Model, Category, Fuel_Type, Typ, Start_Year, End_Year, HSN_TSN
        ''')
        self.conn.commit()

    def check_and_reset_progress(self):
        self.cursor.execute("SELECT value FROM system_metadata WHERE key='last_run_time'")
        row_time = self.cursor.fetchone()
        self.cursor.execute("SELECT value FROM system_metadata WHERE key='scrape_status'")
        row_status = self.cursor.fetchone()

        now = datetime.now()
        needs_reset = False

        if not row_time or not row_status:
            needs_reset = True
        else:
            last_time = datetime.fromisoformat(row_time[0])
            status = row_status[0]
            
            if status == 'completed':
                print("✅ [系統] 上次掃描已全部完成，本次將從頭開始新一輪掃描！")
                needs_reset = True
            elif (now - last_time).days >= 7:
                print("⚠️ [系統] 距離上次執行已超過 7 天，進度重置，從頭開始！")
                needs_reset = True
            else:
                print(f"📦 [系統] 偵測到中斷進度 (上次執行: {last_time.strftime('%Y-%m-%d %H:%M')})，自動接續掃描...")

        if needs_reset:
            self.cursor.execute("DELETE FROM model_progress")
            self.cursor.execute("DELETE FROM system_metadata WHERE key LIKE 'brand_done_%'")
            self.cursor.execute("REPLACE INTO system_metadata (key, value) VALUES ('scrape_status', 'in_progress')")
        
        self.cursor.execute("REPLACE INTO system_metadata (key, value) VALUES ('last_run_time', ?)", (now.isoformat(),))
        self.conn.commit()

    def mark_completed(self):
        self.cursor.execute("REPLACE INTO system_metadata (key, value) VALUES ('scrape_status', 'completed')")
        self.conn.commit()

    def add_to_batch(self, record: dict):
        self.batch.append(record)
        if len(self.batch) >= Config.BATCH_SIZE:
            self.flush()

    def flush(self):
        if not self.batch: return
        changelog_entries = []
        for r in self.batch:
            brand = r.get('Brand', 'N/A')
            model = r.get('Model', 'N/A')
            cat = r.get('Category', 'N/A')
            fuel = r.get('Fuel_Type', 'N/A')
            typ = r.get('Typ', 'N/A')
            sy = r.get('Start_Year', 'N/A')
            ey = r.get('End_Year', 'N/A')
            hsn = r.get('HSN_TSN', 'N/A')

            self.cursor.execute('''
                SELECT rowid, HSN_TSN FROM car_catalog
                WHERE Brand=? AND Model=? AND Category=? AND Fuel_Type=? AND Typ=?
                  AND Start_Year=? AND End_Year=?
            ''', (brand, model, cat, fuel, typ, sy, ey))
            existing = self.cursor.fetchone()

            if existing:
                rowid, old_hsn = existing
                if old_hsn != hsn:
                    if hsn != 'N/A':
                        self.cursor.execute('UPDATE car_catalog SET HSN_TSN=? WHERE rowid=?', (hsn, rowid))
                        r['_hsn_old'] = old_hsn
                        changelog_entries.append(('UPDATE', r))
                    else:
                        pass
                else:
                    pass
            else:
                self.cursor.execute('''
                    INSERT INTO car_catalog
                    (Brand, Model, Category, Fuel_Type, Typ, Start_Year, End_Year, HSN_TSN)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (brand, model, cat, fuel, typ, sy, ey, hsn))
                r['_hsn_old'] = ''
                changelog_entries.append(('INSERT', r))
        self.conn.commit()
        if changelog_entries:
            self._log_changelog(changelog_entries)
        self.batch.clear()

    def get_progress(self, brand: str, model: str):
        self.cursor.execute('SELECT variant_count FROM model_progress WHERE Brand=? AND Model=?', (brand, model))
        row = self.cursor.fetchone()
        return row[0] if row else None

    def update_progress(self, brand: str, model: str, count: int):
        self.cursor.execute('''
            INSERT OR REPLACE INTO model_progress (Brand, Model, variant_count, last_scraped)
            VALUES (?, ?, ?, ?)
        ''', (brand, model, count, datetime.now().isoformat()))
        self.conn.commit()

    def mark_brand_done(self, brand: str):
        self.cursor.execute("REPLACE INTO system_metadata (key, value) VALUES (?, 'true')",
                           (f'brand_done_{brand}',))
        self.conn.commit()

    def is_brand_done(self, brand: str):
        self.cursor.execute("SELECT value FROM system_metadata WHERE key=?", (f'brand_done_{brand}',))
        row = self.cursor.fetchone()
        return row is not None and row[0] == 'true'

    def export_brand_csv(self, brand: str):
        self.flush()
        try:
            df = pd.read_sql_query("SELECT * FROM view_car_catalog WHERE Brand = ?", self.conn, params=(brand,))
            if not df.empty:
                path = os.path.join(Config.CSV_DIR, f"{brand}.csv")
                df.to_csv(path, index=False, encoding='utf-8-sig')
                return len(df)
        except Exception: pass
        return 0

    def get_stats(self):
        try:
            df = pd.read_sql_query('''
                SELECT Brand, Fuel_Type, Category, COUNT(DISTINCT Model) as Models, COUNT(*) as Rows
                FROM view_car_catalog GROUP BY Brand, Fuel_Type, Category ORDER BY Brand, Fuel_Type
            ''', self.conn)
            return df
        except Exception: return pd.DataFrame()

    def get_all_data(self):
        self.flush()
        try:
            return pd.read_sql_query("SELECT * FROM view_car_catalog ORDER BY Brand, Model, Fuel_Type, Typ", self.conn)
        except Exception:
            return pd.DataFrame()

    def close(self):
        self.flush()
        self.conn.close()


class SnapshotManager:
    SNAPSHOT_FILE = Config.SNAPSHOT_FILE
    DIFF_FILE = Config.DIFF_FILE
    KEY_COLS = ['Brand', 'Model', 'Fuel_Type', 'Typ']
    ALL_COLS = ['Brand', 'Model', 'Category', 'Fuel_Type', 'Typ', 'Start_Year', 'End_Year', 'HSN_TSN']

    RED_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    GREEN_FILL = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    GRAY_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_FONT = Font(color='FFFFFF', bold=True)
    COMPARE_FIELDS = ['Category', 'Start_Year', 'End_Year', 'HSN_TSN']

    @staticmethod
    def load_snapshot():
        try:
            df = pd.read_excel(SnapshotManager.SNAPSHOT_FILE, engine='openpyxl')
            print(f"  [Snapshot] 載入上次快照 ({len(df)} 筆)")
            return df
        except Exception:
            return None

    @staticmethod
    def save_snapshot(df):
        try:
            today = datetime.now().strftime('%Y-%m-%d')
            if os.path.exists(SnapshotManager.SNAPSHOT_FILE):
                bak = SnapshotManager.SNAPSHOT_FILE.replace('.xlsx', f'_{today}.xlsx')
                if not os.path.exists(bak):
                    os.rename(SnapshotManager.SNAPSHOT_FILE, bak)
            out = df[SnapshotManager.ALL_COLS] if all(c in df.columns for c in SnapshotManager.ALL_COLS) else df
            out.to_excel(SnapshotManager.SNAPSHOT_FILE, index=False, engine='openpyxl')
        except Exception as e:
            print(f"  [Snapshot] 儲存失敗: {e}")

    @staticmethod
    def generate_diff_report(old_df, new_df):
        if old_df is None or old_df.empty:
            SnapshotManager.save_snapshot(new_df)
            print("  [Diff] 首次執行，已建立基準快照")
            return

        for col in SnapshotManager.ALL_COLS:
            if col in old_df.columns:
                old_df[col] = old_df[col].apply(lambda x: '' if pd.isna(x) else str(x).strip())
            if col in new_df.columns:
                new_df[col] = new_df[col].apply(lambda x: '' if pd.isna(x) else str(x).strip())

        old_df.replace(to_replace=r'^(N/A|n/a|NA|na|nan|NaN|Nan|null|None)$', value='', regex=True, inplace=True)
        new_df.replace(to_replace=r'^(N/A|n/a|NA|na|nan|NaN|Nan|null|None)$', value='', regex=True, inplace=True)

        missing = [c for c in SnapshotManager.KEY_COLS if c not in old_df.columns or c not in new_df.columns]
        if missing:
            print(f"  [Diff] 缺少關鍵欄位 {missing}，跳過差異比較")
            SnapshotManager.save_snapshot(new_df)
            return

        old_df['_key'] = old_df[SnapshotManager.KEY_COLS].agg('|'.join, axis=1)
        new_df['_key'] = new_df[SnapshotManager.KEY_COLS].agg('|'.join, axis=1)

        old_df = old_df.drop_duplicates(subset='_key', keep='last')
        new_df = new_df.drop_duplicates(subset='_key', keep='last')

        old_keys = set(old_df['_key'])
        new_keys = set(new_df['_key'])

        new_records = new_df[~new_df['_key'].isin(old_keys)].copy()
        removed_records = old_df[~old_df['_key'].isin(new_keys)].copy()

        common_keys = old_keys & new_keys
        old_idx = old_df[old_df['_key'].isin(common_keys)].set_index('_key')
        new_idx = new_df[new_df['_key'].isin(common_keys)].set_index('_key')

        compare_cols = [c for c in SnapshotManager.ALL_COLS if c not in SnapshotManager.KEY_COLS]

        modified = []
        for key in common_keys:
            old_row = old_idx.loc[key]
            new_row = new_idx.loc[key]
            changes = {}
            for col in compare_cols:
                ov = str(old_row.get(col, ''))
                nv = str(new_row.get(col, ''))
                if ov != nv:
                    changes[col] = (ov, nv)
            if changes:
                modified.append((old_row.to_dict(), new_row.to_dict(), changes))

        has_new = not new_records.empty
        has_mod = len(modified) > 0
        has_del = not removed_records.empty

        if not (has_new or has_mod or has_del):
            print("  [Diff] ✅ 無任何差異")
            SnapshotManager.save_snapshot(new_df)
            return

        today = datetime.now().strftime('%Y-%m-%d')
        if os.path.exists(SnapshotManager.DIFF_FILE):
            wb = load_workbook(SnapshotManager.DIFF_FILE)
            if today in wb.sheetnames:
                ws = wb[today]
                row = ws.max_row + 2
                ws.cell(row=row - 1, column=1, value=f'--- {datetime.now().strftime("%H:%M")} ---')
            else:
                ws = wb.create_sheet(today, 0)
                row = 1
        else:
            wb = Workbook()
            wb.remove(wb.active)
            ws = wb.create_sheet(today, 0)
            row = 1

        if row == 1:
            headers = ['狀態'] + SnapshotManager.KEY_COLS
            for f in SnapshotManager.COMPARE_FIELDS:
                headers.append(f'{f}(O)')
                headers.append(f'{f}(N)')
            SnapshotManager._write_header(ws, headers)
            row = 2

        for _, old_row in removed_records.iterrows():
            SnapshotManager._write_status_row(ws, row, '刪除', old_row.to_dict(), None, SnapshotManager.RED_FILL, None)
            row += 1

        for old_data, new_data, changes in modified:
            SnapshotManager._write_status_row(ws, row, '修改', old_data, new_data, SnapshotManager.RED_FILL, SnapshotManager.GREEN_FILL, changes)
            row += 1

        for _, new_row in new_records.iterrows():
            SnapshotManager._write_status_row(ws, row, '新增', None, new_row.to_dict(), None, SnapshotManager.GREEN_FILL)
            row += 1

        wb.save(SnapshotManager.DIFF_FILE)
        print(f"\n  [Diff] 📊 差異報告: {SnapshotManager.DIFF_FILE} (分頁: {today})")
        print(f"    ├─ 新增: {len(new_records)} 筆")
        print(f"    ├─ 修改: {len(modified)} 筆")
        print(f"    └─ 刪除: {len(removed_records)} 筆")

        SnapshotManager.save_snapshot(new_df)

    @staticmethod
    def _write_header(ws, headers):
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = SnapshotManager.HEADER_FILL
            cell.font = SnapshotManager.HEADER_FONT

    @staticmethod
    def _write_status_row(ws, row, status, old_data, new_data, old_fill, new_fill, changes=None):
        ws.cell(row=row, column=1, value=status)
        for ci, col in enumerate(SnapshotManager.KEY_COLS, 2):
            val = (new_data or old_data or {}).get(col, '')
            ws.cell(row=row, column=ci, value=val)

        for fi, f in enumerate(SnapshotManager.COMPARE_FIELDS):
            col_o = 2 + len(SnapshotManager.KEY_COLS) + fi * 2
            col_n = col_o + 1
            ov = (old_data or {}).get(f, '')
            nv = (new_data or {}).get(f, '')
            ws.cell(row=row, column=col_o, value=ov if ov else '')
            ws.cell(row=row, column=col_n, value=nv if nv else '')
            is_changed = changes is None or f in changes
            if old_fill and ov and is_changed: ws.cell(row=row, column=col_o).fill = old_fill
            if new_fill and nv and is_changed: ws.cell(row=row, column=col_n).fill = new_fill


async def smart_delay(success=True):
    delay = random.uniform(Config.DELAY_MIN, Config.DELAY_MAX) if success else random.uniform(2.5, 4.5)
    if random.random() < 0.12: delay += random.uniform(1.0, 2.5)
    await asyncio.sleep(max(0.5, min(delay, 6.0)))

async def dismiss_cookie(page):
    try:
        iframe = page.frame_locator('iframe[id^="sp_message_iframe"]')
        btn = iframe.get_by_role("button", name="Alle akzeptieren")
        await btn.click(timeout=5000)
    except Exception: pass

def extract_date_range(text):
    if not text: return "N/A"
    match = re.search(r'(\d{2}/\d{4})\s*[–-]\s*(\d{2}/\d{4})', text)
    if match: return f"{match.group(1)} - {match.group(2)}"
    match = re.search(r'seit\s+(\d{2}/\d{4})', text)
    if match: return f"seit {match.group(1)}"
    return "N/A"

class AutoBildScraper:
    def __init__(self, test_mode=False, target_brand=None, diff=True):
        self.test_mode = test_mode
        self.target_brand = target_brand.upper() if target_brand else None
        self.diff = diff
        self.db = DatabaseManager()
        self.start_time = time.time()
        self.api_hsn_tsn = None
        self.api_captured = False
        self.is_interrupted = False
        self._setup_signal_handlers()

    def _setup_signal_handlers(self):
        def signal_handler(signum, frame):
            print(f"\n\n[⚠️ 警告] 偵測到人為中斷 (Ctrl+C 或關閉)，系統正在安全儲存進度...")
            self.is_interrupted = True
            
        try:
            signal.signal(signal.SIGINT, signal_handler)
            signal.signal(signal.SIGTERM, signal_handler)
        except Exception:
            pass 

    def should_stop(self):
        if self.is_interrupted:
            return True
        if not Config.ENABLE_RUNTIME_LIMIT:
            return False
        elapsed = (time.time() - self.start_time) / 3600
        if elapsed >= Config.MAX_RUNTIME_HOURS:
            print(f"\n[Timeout] 觸發超時保護 (>{Config.MAX_RUNTIME_HOURS} hours), 準備安全停止...")
            return True
        return False

    async def handle_api_response(self, response):
        try:
            url = response.url.lower()
            if response.status != 200: return
            ct = response.headers.get('content-type', '')
            if 'json' not in ct and not url.endswith('.json'): return
            if '/api/' not in url: return

            try:
                data = await response.json()
            except Exception:
                return
            if not isinstance(data, (dict, list)): return

            data_str = str(data)
            if 'tcert' in data_str.lower():
                print(f'  [API] tcert FOUND in: {response.url[:180]}')
                self.api_hsn_tsn = data
                self.api_captured = True
            elif url.startswith('https://www.autobild.de/api'):
                print(f'  [API] {response.url[:180]}')
        except Exception:
            pass
        except Exception: pass

    async def collect_brand_urls(self, page):
        print("\n[Step 1] Collecting brand URLs...")
        for attempt in range(3):
            try:
                await page.goto(Config.CATALOG_URL, timeout=90000, wait_until="domcontentloaded")
                break
            except Exception as e:
                if attempt < 2:
                    wait = (attempt + 1) * 5
                    print(f"    [Retry] 品牌列表載入失敗 ({e.__class__.__name__}), {wait}秒後重試 ({attempt+1}/3)...")
                    await asyncio.sleep(wait)
                else:
                    raise
        await dismiss_cookie(page)

        for scroll_y in range(0, 3000, 600):
            await page.evaluate(f"window.scrollTo(0, {scroll_y})")
            await asyncio.sleep(0.5)
        await smart_delay()

        raw_links = await page.evaluate(r'''() => {
            return Array.from(document.querySelectorAll('a[href*="/marken-modelle/"]'))
                .map(a => ({ href: a.getAttribute('href'), text: (a.textContent || '').trim() }))
                .filter(l => l.href && l.text.length > 0);
        }''')

        brand_urls = []
        seen = set()
        for item in raw_links:
            href = item['href']
            if not href or '/marken-modelle/' not in href: continue

            full = urljoin(Config.BASE_URL, href).split('#')[0].split('?')[0]
            full = full.rstrip('/') + '/'
            parts = full.replace(Config.BASE_URL, '').strip('/').split('/')
            
            if len(parts) == 2 and parts[0] == 'marken-modelle' and parts[1]:
                if full not in seen:
                    seen.add(full)
                    brand_urls.append(full)

        brand_urls.sort(reverse=True)
        if self.target_brand:
            brand_urls = [u for u in brand_urls if f'/{self.target_brand.lower()}/' in u]
        elif self.test_mode:
            brand_urls = brand_urls[:2]

        print(f"   Found {len(brand_urls)} brands")
        return brand_urls

    async def collect_model_urls(self, page, brand_url):
        brand_name = brand_url.strip('/').split('/')[-1].upper()
        print(f"\n> Entering brand: {brand_name}")
        for attempt in range(3):
            try:
                await page.goto(brand_url, timeout=90000, wait_until="domcontentloaded")
                break
            except Exception as e:
                if 'ERR_NETWORK_CHANGED' in str(e) and attempt < 2:
                    wait = (attempt + 1) * 3
                    print(f"    [Retry] 網路不穩 ({e.__class__.__name__}), {wait}秒後重試 ({attempt+1}/3)...")
                    await asyncio.sleep(wait)
                else:
                    raise
        await dismiss_cookie(page)
        await smart_delay()

        await page.evaluate(r'''() => {
            const allEls = Array.from(document.querySelectorAll('a, button, div, span, li, h2, h3'));
            const btn = allEls.find(el => (el.textContent || '').replace(/\s+/g, '').trim().toUpperCase() === 'ALLEMODELLE');
            if (btn) btn.click();
        }''')
        await asyncio.sleep(2.5)

        raw_links = await page.evaluate(r'''() => {
            return Array.from(document.querySelectorAll('a[href*="/marken-modelle/"]'))
                .map(a => a.getAttribute('href')).filter(h => h);
        }''')

        models = []
        seen = set()
        brand_path = brand_url.rstrip('/')
        for href in raw_links:
            full = urljoin(Config.BASE_URL, href).split('#')[0].split('?')[0]
            if full.startswith(brand_path + '/') and full != brand_path + '/':
                segments = full.replace(Config.BASE_URL, '').strip('/').split('/')
                if len(segments) == 3:
                    clean = full.rstrip('/') + '/'
                    if clean not in seen:
                        seen.add(clean)
                        models.append(clean)

        models = list(dict.fromkeys(models))
        if self.test_mode: models = models[:2]

        print(f"   Found {len(models)} models")
        return brand_name, models

    async def expand_all_variants(self, page):
        for _ in range(20):
            clicked = await page.evaluate(r'''() => {
                const btns = document.querySelectorAll('.vv__fuelType-dataFooterToggle');
                for (const btn of btns) {
                    if (btn.offsetParent !== null) {
                        btn.click(); return true;
                    }
                }
                return false;
            }''')
            if not clicked: break
            await asyncio.sleep(1.5)

    async def extract_page_data(self, page):
        data = await page.evaluate(r'''() => {
            const result = {
                headerTitle: 'N/A', headerSubtitle: 'N/A', fuelTypes: [], variants: [],
                editorialTable: null, jsonBuildingPeriod: null, jsonBodyType: null
            };

            const titleEl = document.querySelector('.vv__header-text-title');
            if (titleEl) result.headerTitle = titleEl.textContent.trim();
            const subtitleEl = document.querySelector('.vv__header-text-subtitle');
            if (subtitleEl) result.headerSubtitle = subtitleEl.textContent.trim();

            const fuelSections = document.querySelectorAll('.vv__fuelType');
            fuelSections.forEach(section => {
                const badge = section.querySelector('.fuelTypeBadge');
                const fuelType = badge ? badge.textContent.trim() : 'N/A';

                const rows = section.querySelectorAll('.vv__fuelType-dataBodyLine');
                rows.forEach(row => {
                    const cells = row.querySelectorAll('div');
                    if (cells.length >= 5) {
                        const rawVariant = cells[0].textContent.trim();
                        const match = rawVariant.match(/^(.+?)\s+(\d{2}\/\d{4})\s*[–-]\s*(\d{2}\/\d{4})/);
                        const variantName = match ? match[1].trim() : rawVariant;
                        const dateRange = match ? match[2] + ' - ' + match[3] : (rawVariant.match(/seit\s+\d{2}\/\d{4}/) || [''])[0];

                        result.variants.push({
                            fuelType: fuelType, name: variantName, dateRange: dateRange,
                            power: cells[1] ? cells[1].textContent.trim() : 'N/A',
                            acceleration: cells[2] ? cells[2].textContent.trim() : 'N/A',
                            consumption: cells[3] ? cells[3].textContent.trim() : 'N/A',
                            price: cells[4] ? cells[4].textContent.trim() : 'N/A'
                        });
                    }
                });
            });

            try {
                const ctxEl = document.querySelector('#vike_pageContext');
                if (ctxEl) {
                    const ctx = JSON.parse(ctxEl.textContent);
                    const mg = ctx.irContent && ctx.irContent.modelGeneration;
                    if (mg) {
                        if (mg.buildingPeriod) {
                            const from = mg.buildingPeriod.fromYear || '';
                            const till = mg.buildingPeriod.tillYear || '';
                            result.jsonBuildingPeriod = till ? from + ' - ' + till : 'seit ' + from;
                        }
                        if (mg.constructionTypeImages && mg.constructionTypeImages[0]) {
                            result.jsonBodyType = mg.constructionTypeImages[0].type || null;
                        }
                    }
                }
            } catch(e) {}
            return result;
        }''')
        return data

    async def try_extract_hsn_tsn_from_overlay(self, page, variant_index):
        self.api_hsn_tsn = None
        self.api_captured = False

        try:
            clicked = await page.evaluate(r'''(idx) => {
                const rows = document.querySelectorAll('.vv__fuelType-dataBodyLine');
                if (idx < rows.length) {
                    const link = rows[idx].querySelector('.vv__fuelType-dataBodyLineLink');
                    if (link) { link.click(); return true; }
                }
                return false;
            }''', variant_index)

            print(f"    [DEBUG overlay {variant_index}] clicked={clicked}")
            if not clicked: return "N/A"
            await asyncio.sleep(Config.API_WAIT_TIMEOUT)

            print(f"    [DEBUG overlay {variant_index}] api_captured={self.api_captured}, api_hsn_tsn={'YES' if self.api_hsn_tsn else 'None'}")

            result = "N/A"
            if self.api_hsn_tsn:
                data_str = str(self.api_hsn_tsn)
                print(f"    [DEBUG] api_hsn_tsn keys={list(self.api_hsn_tsn.keys()) if isinstance(self.api_hsn_tsn, dict) else type(self.api_hsn_tsn)} data={data_str[:300]}")
                hsn, tsn = self._find_hsn_tsn_pair(self.api_hsn_tsn)
                if hsn and tsn: result = f"{hsn}/{tsn}"
                elif hsn and not tsn: result = hsn

            if result == "N/A":
                evaluate_result = await page.evaluate(r'''() => {
                    const overlay = document.querySelector('.vvp') || document.querySelector('.sectionOverlay') || document.querySelector('[class*="overlay"]');
                    if (!overlay) return JSON.stringify({hsn: null, text: '__NO_OVERLAY__'});
                    const txt = overlay.innerText || overlay.textContent || '';
                    const m = txt.match(/HSN\/TSN\s*Schlüsselnummern[\s\S]{0,80}/i);
                    if (m) {
                        const pairs = [];
                        const re = /(\d{4})\s*[\/\-]\s*([A-Z0-9]{2,6})/gi;
                        let mm;
                        while ((mm = re.exec(m[0])) !== null) pairs.push(mm[1] + '/' + mm[2]);
                        if (pairs.length) return JSON.stringify({hsn: pairs.join(', '), text: txt.slice(0,500)});
                    }
                    const hsnL = txt.match(/HSN[:\s]*(\d{4})/i);
                    const tsnL = txt.match(/TSN[:\s]*([A-Z0-9]{2,6})/i);
                    if (hsnL && tsnL) return JSON.stringify({hsn: hsnL[1] + '/' + tsnL[1], text: txt.slice(0,500)});
                    const gen = txt.match(/(\d{4})\s*[\/\-]\s*([A-Z0-9]{2,6})/);
                    const h = gen ? gen[1] + '/' + gen[2] : null;
                    return JSON.stringify({hsn: h, text: txt.slice(0,500)});
                }''')
                dom_data = json.loads(evaluate_result)
                print(f"    [DEBUG overlay {variant_index}] dom_hsn={dom_data.get('hsn')}, overlay_text={dom_data.get('text','')[:200]}")
                if dom_data.get('hsn'): result = dom_data['hsn']

            await page.evaluate("const btn = document.querySelector('.sectionOverlay__buttonClose'); if (btn) btn.click();")
            await asyncio.sleep(0.5)
            print(f"    [DEBUG overlay {variant_index}] result={result}")
            return result

        except Exception as e:
            print(f"    [DEBUG overlay {variant_index}] EXCEPTION: {e}")
            try:
                await page.evaluate("const btn = document.querySelector('.sectionOverlay__buttonClose'); if (btn) btn.click();")
                await asyncio.sleep(0.5)
            except Exception: pass
        return "N/A"

    def _find_hsn_tsn_pair(self, data, depth=0):
        if depth > 10: return None, None
        if isinstance(data, dict):
            tcert_key = next((k for k in data if 'tcert' in str(k).lower()), None)
            if tcert_key:
                tcert_list = data[tcert_key]
                pairs = []
                if isinstance(tcert_list, list):
                    for entry in tcert_list:
                        if isinstance(entry, dict) and 'Num' in entry and 'Num2' in entry:
                            hv = str(entry.get('Num', '') or '').strip()
                            tv = str(entry.get('Num2', '') or '').strip()
                            if hv and tv and hv != '-' and tv != '-':
                                hsn_clean = re.sub(r'\D', '', hv)[:4]
                                tsn_clean = re.sub(r'[^A-Z0-9]', '', tv.upper())[:6]
                                if hsn_clean and tsn_clean:
                                    pairs.append(f"{hsn_clean}/{tsn_clean}")
                    if pairs:
                        return ', '.join(pairs), None
                elif isinstance(tcert_list, dict) and 'Num' in tcert_list and 'Num2' in tcert_list:
                    hv = str(tcert_list.get('Num', '') or '').strip()
                    tv = str(tcert_list.get('Num2', '') or '').strip()
                    if hv and tv and hv != '-' and tv != '-':
                        hsn_clean = re.sub(r'\D', '', hv)[:4]
                        tsn_clean = re.sub(r'[^A-Z0-9]', '', tv.upper())[:6]
                        if hsn_clean and tsn_clean:
                            return f"{hsn_clean}/{tsn_clean}", None
            hsn_k = next((k for k in data if 'hsn' in str(k).lower()), None)
            tsn_k = next((k for k in data if 'tsn' in str(k).lower()), None)
            if hsn_k and tsn_k:
                hv = str(data[hsn_k]).strip() if data[hsn_k] else ''
                tv = str(data[tsn_k]).strip() if data[tsn_k] else ''
                if hv and tv and hv != '-' and tv != '-':
                    hsn_clean = re.sub(r'\D', '', hv)[:4]
                    tsn_clean = re.sub(r'[^A-Z0-9]', '', tv.upper())[:6]
                    if hsn_clean and tsn_clean: return hsn_clean, tsn_clean
            for v in data.values():
                h, t = self._find_hsn_tsn_pair(v, depth + 1)
                if h and t: return h, t
                if h and not t: return h, None
        elif isinstance(data, list):
            for item in data:
                h, t = self._find_hsn_tsn_pair(item, depth + 1)
                if h and t: return h, t
        return None, None

    def build_records(self, brand, model, page_data, fuel_type_filter=None, fallback=False):
        records = []
        body_type = page_data.get('jsonBodyType') or 'N/A'
        
        body_map = {
            'fliessheck': '掀背車',
            'stufenheck': '斜背車',
            'stlheck_fliessheck': '斜掀背車',
            'steilheck_fliessheck': '高背掀背車',
            'limousine': '轎車',
            'suv': '休旅車',
            'cabrio': '敞篷車',
            'roadster': '敞篷跑車',
            'coupe': '雙門跑車',
            'kombi': '旅行車',
            'kasten': '貨車',
            'van': '廂型車',
            'bus': '巴士',
            'select': '精選型',
            'pritsche': '皮卡',
            'geländewagen': '越野車',
            'gelaendewagen': '越野車',
            'kleinwagen': '小型車'
        }
        if body_type.lower() in body_map:
            body_type = body_map[body_type.lower()]

        fuel_map = {
            'benzin': '汽油',
            'diesel': '柴油',
            'elektro': '電動',
            'elektrischer strom': '電力',
            'benzin/hybrid': '油電混合',
            'benzin/elektro': '油電混合',
            'benzin/elektro-plugin': '插電式油電混合',
            'benzin/elektro-plug': '插電式油電混合',
            'benzin/gas': '汽油/天然氣',
            'benzin/alkohol': '汽油/酒精混合',
            'erdgas': '天然氣',
            'autogas': '液化石油氣',
            'plug-in-hybrid': '插電式油電混合',
            'wasserstoff': '氫燃料'
        }

        for v in page_data.get('variants', []):
            fuel = v.get('fuelType', 'N/A')
            if fuel_type_filter and fuel != fuel_type_filter:
                continue
            
            fuel_zh = fuel_map.get(fuel.lower(), fuel)

            v_date = v.get('dateRange', '')
            start_year, end_year = 'N/A', 'N/A'
            if v_date:
                parts = re.split(r'\s*[–-]\s*', v_date.replace('seit ', ''))
                if len(parts) == 2:
                    sy = parts[0].strip().split('/')
                    ey = parts[1].strip().split('/')
                    start_year = sy[1] if len(sy) > 1 else sy[0]
                    end_year = ey[1] if len(ey) > 1 else ey[0]
                elif len(parts) == 1:
                    sy = parts[0].strip().split('/')
                    start_year = sy[1] if len(sy) > 1 else sy[0]
                    end_year = '至今'

            records.append({
                'Brand': brand, 
                'Model': model, 
                'Category': body_type, 
                'Fuel_Type': fuel_zh,
                'Typ': v.get('name', 'N/A'), 
                'Start_Year': start_year, 
                'End_Year': end_year, 
                'HSN_TSN': 'N/A'
            })
            
        if not records and fallback:
            period = page_data.get('jsonBuildingPeriod') or 'N/A'
            start_year, end_year = 'N/A', 'N/A'
            if period != 'N/A':
                parts = re.split(r'\s*[–-]\s*', period.replace('seit ', ''))
                if len(parts) == 2:
                    start_year, end_year = parts[0].strip(), parts[1].strip()
                elif len(parts) == 1:
                    start_year, end_year = parts[0].strip(), '至今'
            records.append({
                'Brand': brand, 'Model': model, 'Category': body_type,
                'Fuel_Type': 'N/A', 'Typ': f'{model} ({period})' if period != 'N/A' else model,
                'Start_Year': start_year, 'End_Year': end_year, 'HSN_TSN': 'N/A'
            })
            
        return records

    async def process_model(self, page, brand, model_url):
        model_name = model_url.strip('/').split('/')[-1].replace('-', ' ').title()
        
        if self.should_stop():
            self.db.flush()
            return False

        await page.goto(model_url, timeout=60000, wait_until="domcontentloaded")
        await dismiss_cookie(page)

        for scroll_y in range(0, 2000, 400):
            await page.evaluate(f"window.scrollTo(0, {scroll_y})")
            await asyncio.sleep(0.3)
        await asyncio.sleep(1.0)

        has_variants = await page.evaluate("() => document.querySelectorAll('.vv__fuelType-dataBodyLine').length > 0")

        if not has_variants:
            page_data = await self.extract_page_data(page)
            records = self.build_records(brand, model_name, page_data)
            if not records:
                records = self.build_records(brand, model_name, page_data, fallback=True)
            if records:
                saved = self.db.get_progress(brand, model_name)
                if saved is not None and saved == len(records) and not self.test_mode:
                    print(f"  [Skip] [{model_name}] (已完成 {saved} 筆)")
                    return True

                for r in records: self.db.add_to_batch(r)
                self.db.flush()
                if not self.should_stop():
                    self.db.update_progress(brand, model_name, len(records))
                    csv_count = self.db.export_brand_csv(brand)
                    print(f"    [CSV] 已更新 {brand}.csv ({csv_count} 筆)")
                print(f"  [OK] [{model_name}] 單一規格: {len(records)} 筆")
            return True

        await self.expand_all_variants(page)
        page_data = await self.extract_page_data(page)

        variant_count = len(page_data.get('variants', []))
        saved_count = self.db.get_progress(brand, model_name)

        if saved_count is not None and saved_count == variant_count and not self.test_mode:
            print(f"  [Skip] [{model_name}] (已知 {saved_count} 款)")
            return True

        print(f"  -> [{brand} - {model_name}] 共 {variant_count} 款，開始擷取...")
        records = self.build_records(brand, model_name, page_data)
        if self.test_mode: records = records[:6]

        for i, record in enumerate(records):
            if self.should_stop(): break
            
            sys.stdout.write(f"\r      [{i+1}/{len(records)}] {record['Fuel_Type'][:8]} - {record['Typ'][:25]}...")
            sys.stdout.flush()

            if record['HSN_TSN'] == 'N/A':
                hsn = await self.try_extract_hsn_tsn_from_overlay(page, i)
                if hsn != 'N/A': record['HSN_TSN'] = hsn

            self.db.add_to_batch(record)

        print()
        
        if not self.should_stop():
            self.db.flush()
            self.db.update_progress(brand, model_name, variant_count)
            csv_count = self.db.export_brand_csv(brand)
            print(f"    [CSV] 已更新 {brand}.csv ({csv_count} 筆)")
        else:
            self.db.flush()
            return False
            
        return True

    async def run(self):
        print("\n" + "=" * 50)
        print("  AutoBild Scraper v11.0 - VS Code 本地完美中斷存檔版")
        print(f"  Mode: {'Test' if self.test_mode else 'Full'}")
        if self.target_brand: print(f"  Target Brand: {self.target_brand}")
        print(f"  Timeout Protection: {'enabled' if Config.ENABLE_RUNTIME_LIMIT else 'disabled (run until complete)'}")
        print("=" * 50 + "\n")

        db_bak = Config.DB_FILE.replace('.db', '_backup.db')
        if os.path.exists(Config.DB_FILE):
            try:
                conn = sqlite3.connect(Config.DB_FILE)
                count = conn.execute("SELECT COUNT(*) FROM car_catalog").fetchone()[0]
                conn.close()
            except Exception:
                count = 0
            if count > 0:
                try:
                    import shutil
                    shutil.copy2(Config.DB_FILE, db_bak)
                    print(f"  [Backup] DB 備份: {db_bak} ({count} 筆)")
                except Exception as e:
                    print(f"  [Backup] DB 備份失敗: {e}")
            else:
                print("  [Backup] DB 為空，略過備份")

        old_snapshot = SnapshotManager.load_snapshot()

        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            context = await browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                viewport={'width': 1440, 'height': 900}
            )
            context.on("response", self.handle_api_response)
            page = await context.new_page()

            try:
                brand_urls = await self.collect_brand_urls(page)
                total_brands = len(brand_urls)
                
                for b_idx, b_url in enumerate(brand_urls):
                    if self.should_stop(): break

                    brand_name = b_url.strip('/').split('/')[-1].upper()

                    if self.db.is_brand_done(brand_name) and not self.test_mode:
                        print(f"\n  [Skip] [{brand_name}] 已完成，跳過")
                        continue

                    brand_name, model_urls = await self.collect_model_urls(page, b_url)

                    parsed_ok = True
                    for m_idx, m_url in enumerate(model_urls):
                        if self.should_stop(): break
                        try:
                            ok = await self.process_model(page, brand_name, m_url)
                            if not ok: parsed_ok = False
                        except Exception as e:
                            print(f"    [Error] Model processing failed: {e}")
                            parsed_ok = False

                    if not self.should_stop():
                        csv_count = self.db.export_brand_csv(brand_name)
                        elapsed = (time.time() - self.start_time) / 60
                        print(f"\n   [Export] {brand_name} done ({csv_count} records) [{b_idx+1}/{total_brands} brands]")
                        if parsed_ok:
                            self.db.mark_brand_done(brand_name)

                if not self.should_stop() and not self.test_mode and not self.target_brand:
                    self.db.mark_completed()
                    print("\n🎉 全站掃描完畢！已標記為完成，下次啟動將自動重新掃描。")

            except Exception as e:
                print(f"\n[Error] Execution interrupted: {e}")
            finally:
                self.db.flush()
                await browser.close()

        new_snapshot = self.db.get_all_data()
        if self.diff and not new_snapshot.empty:
            SnapshotManager.generate_diff_report(old_snapshot, new_snapshot)

        self.db.close()
        elapsed = (time.time() - self.start_time) / 60
        print(f"\n[Complete] 任務結束！總耗時: {elapsed:.1f} 分鐘")
        print(f"[Database] {Config.DB_FILE}")


def show_status():
    if not os.path.exists(Config.DB_FILE):
        print("[Error] 找不到資料庫")
        return
    conn = sqlite3.connect(Config.DB_FILE)
    try:
        df = pd.read_sql_query('''
            SELECT Brand, Fuel_Type, COUNT(DISTINCT Model) as Models, COUNT(*) as Rows
            FROM view_car_catalog GROUP BY Brand, Fuel_Type ORDER BY Brand, Fuel_Type
        ''', conn)
        if df.empty:
            print("[Info] 資料庫是空的")
        else:
            print("\n[統計資料]")
            print(df.to_string())
            total = pd.read_sql_query("SELECT COUNT(*) as Total FROM view_car_catalog", conn)
            print(f"\n  總筆數: {total.iloc[0]['Total']}")
    except Exception as e: print(f"[Error] {e}")
    finally: conn.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="AutoBild Scraper v11.0")
    parser.add_argument("--test", action="store_true", help="Test mode")
    parser.add_argument("--reset", action="store_true", help="Reset database and snapshot")
    parser.add_argument("--status", action="store_true", help="Show statistics")
    parser.add_argument("--brand", type=str, default=None, help="Scrape specific brand")
    parser.add_argument("--diff", action="store_true", default=True, help="Generate diff report (default: on)")
    parser.add_argument("--no-diff", action="store_false", dest="diff", help="Skip diff report")
    args = parser.parse_args()

    if args.reset:
        for f in [Config.DB_FILE]:
            if os.path.exists(f):
                os.remove(f)
        for f in glob.glob(os.path.join(Config.CSV_DIR, 'autobild_snapshot_*.xlsx')):
            os.remove(f)
        for pat in ['autobild_snapshot*.xlsx', 'autobild_diff*.xlsx']:
            for f in glob.glob(os.path.join(Config.CSV_DIR, pat)):
                os.remove(f)
        print("[OK] 資料庫與快照已重置！")
        sys.exit(0)

    if args.status:
        show_status()
        sys.exit(0)

    scraper = AutoBildScraper(test_mode=args.test, target_brand=args.brand, diff=args.diff)
    asyncio.run(scraper.run())